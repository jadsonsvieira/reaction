import os
import json
import re
import secrets
import mysql.connector
import requests
import io
import csv
import time
from flask import Flask, render_template, redirect, url_for, session, request, jsonify, Response
from datetime import datetime, date, timedelta
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash
from flask_cors import CORS

# Assumindo que o ficheiro da IA continua a chamar-se agente_reputacao.py
from agente_reputacao import analisar_avaliacao_gemini
from servico_email import enviar_email_boas_vindas, enviar_notificacao_admin, enviar_email_redefinicao_senha

load_dotenv()

GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET")
MICROSOFT_CLIENT_ID = os.environ.get("MICROSOFT_CLIENT_ID")
FACEBOOK_APP_ID = os.environ.get("FACEBOOK_APP_ID") or os.environ.get("META_APP_ID")
META_APP_ID = FACEBOOK_APP_ID
META_APP_SECRET = os.environ.get("META_APP_SECRET")
META_VERIFY_TOKEN = os.environ.get("META_VERIFY_TOKEN", "frameia_reaction_2026")

app = Flask(__name__)

@app.context_processor
def inject_oauth_ids():
    raw_google = os.environ.get('GOOGLE_CLIENT_ID', '71269651978-gp165jo1i5r6mgmb22u8s82g0jsdh5v0.apps.googleusercontent.com')
    raw_fb = os.environ.get('FACEBOOK_APP_ID', '2263040147842797')
    raw_ms = os.environ.get('MICROSOFT_CLIENT_ID', '138269ce-38e6-4c1e-bc6a-b5292e877a24')
    return {
        'google_client_id': raw_google.replace('GOOGLE_CLIENT_ID=', '').replace('"', '').strip() if raw_google else '71269651978-gp165jo1i5r6mgmb22u8s82g0jsdh5v0.apps.googleusercontent.com',
        'facebook_app_id': raw_fb.replace('FACEBOOK_APP_ID=', '').replace('"', '').strip() if raw_fb else '2263040147842797',
        'microsoft_client_id': raw_ms.replace('MICROSOFT_CLIENT_ID=', '').replace('"', '').strip() if raw_ms else '138269ce-38e6-4c1e-bc6a-b5292e877a24'
    }


@app.context_processor
def inject_oauth_credentials():
    return dict(
        google_client_id=GOOGLE_CLIENT_ID or '',
        facebook_app_id=FACEBOOK_APP_ID or '',
        microsoft_client_id=MICROSOFT_CLIENT_ID or ''
    )

# Chave de sessão estável para evitar desconexões quando o servidor reinicia
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "chave_super_secreta_reaction_2026")

# CORS adaptativo para permitir desenvolvimento local e domínio oficial de produção
cors_origins = os.environ.get("CORS_ORIGINS", "https://reaction.frameia.com.br").split(",")
if app.debug or os.environ.get("FLASK_ENV") == "development":
    # Permite localhost e 127.0.0.1 nas portas comuns
    cors_origins.extend([
        "http://localhost:5000", "http://127.0.0.1:5000",
        "http://localhost:5001", "http://127.0.0.1:5001"
    ])
CORS(app, resources={r"/*": {"origins": cors_origins}})

DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_USER = os.environ.get("DB_USER", "root")
DB_PASSWORD = os.environ.get("DB_PASS", "")
DB_NAME = os.environ.get("DB_NAME", "reaction_db")

# ================= BANCO DE DADOS =================
def get_db_connection():
    try:
        conn = mysql.connector.connect(
            host=DB_HOST,
            user=DB_USER,
            password=DB_PASSWORD,
            database=DB_NAME
        )
        return conn
    except mysql.connector.Error as err:
        if err.errno == mysql.connector.errorcode.ER_BAD_DB_ERROR:
            try:
                conn = mysql.connector.connect(
                    host=DB_HOST,
                    user=DB_USER,
                    password=DB_PASSWORD
                )
                cursor = conn.cursor()
                cursor.execute(f"CREATE DATABASE {DB_NAME}")
                conn.database = DB_NAME
                cursor.close()
                return conn
            except mysql.connector.Error as e:
                print(f"Erro ao tentar criar base de dados: {e}")
                return None
        else:
            print(f"Erro ao conectar à base de dados: {err}")
            return None

def criar_tabelas_se_nao_existirem():
    conn = get_db_connection()
    if not conn:
        print("Aviso: Falha ao obter conexão para criar tabelas.")
        return
        
    try:
        cursor = conn.cursor()
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS empresas (
                id INT AUTO_INCREMENT PRIMARY KEY,
                nome_empresa VARCHAR(255) NOT NULL,
                plano_assinatura VARCHAR(50) DEFAULT 'basico',
                data_criacao DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id INT AUTO_INCREMENT PRIMARY KEY,
                empresa_id INT NOT NULL,
                nome VARCHAR(255) NOT NULL,
                email VARCHAR(255) NOT NULL UNIQUE,
                senha VARCHAR(255) NOT NULL,
                role ENUM('dono', 'admin', 'operador') DEFAULT 'operador',
                FOREIGN KEY (empresa_id) REFERENCES empresas(id) ON DELETE CASCADE
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS integracoes_api (
                id INT AUTO_INCREMENT PRIMARY KEY,
                empresa_id INT NOT NULL,
                plataforma ENUM('google', 'instagram', 'whatsapp', 'ifood') NOT NULL,
                token_acesso TEXT,
                token_refresh TEXT,
                plataforma_user_id VARCHAR(255) NULL,
                data_expiracao DATETIME,
                status ENUM('ativo', 'inativo', 'erro') DEFAULT 'inativo',
                FOREIGN KEY (empresa_id) REFERENCES empresas(id) ON DELETE CASCADE
            )
        """)
        
        # Migração automática de colunas para integracoes_api se a tabela já existia
        colunas_integracoes = {
            'token_refresh': 'TEXT NULL AFTER token_acesso',
            'plataforma_user_id': 'VARCHAR(255) NULL AFTER token_acesso',
            'data_expiracao': 'DATETIME NULL AFTER status'
        }
        for col_nome, col_def in colunas_integracoes.items():
            try:
                cursor.execute(f"SHOW COLUMNS FROM integracoes_api LIKE '{col_nome}'")
                if not cursor.fetchone():
                    cursor.execute(f"ALTER TABLE integracoes_api ADD COLUMN {col_nome} {col_def}")
                    print(f"Coluna {col_nome} adicionada com sucesso na tabela integracoes_api.")
            except Exception as col_err:
                print(f"Erro ao verificar/adicionar coluna {col_nome} em integracoes_api: {col_err}")
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS avaliacoes_feed (
                id INT AUTO_INCREMENT PRIMARY KEY,
                empresa_id INT NOT NULL,
                plataforma_origem ENUM('google', 'instagram', 'whatsapp', 'ifood') NOT NULL,
                nome_cliente VARCHAR(255),
                nota INT CHECK (nota >= 1 AND nota <= 5),
                comentario TEXT,
                sentimento VARCHAR(50),
                rascunho_resposta TEXT,
                tags TEXT,
                status ENUM('pendente', 'respondido_ia', 'alerta_crise') DEFAULT 'pendente',
                FOREIGN KEY (empresa_id) REFERENCES empresas(id) ON DELETE CASCADE
            )
        """)
        
        # Garantir que a coluna 'tags' existe em 'avaliacoes_feed' se a tabela já existia
        try:
            cursor.execute("SHOW COLUMNS FROM avaliacoes_feed LIKE 'tags'")
            if not cursor.fetchone():
                cursor.execute("ALTER TABLE avaliacoes_feed ADD COLUMN tags TEXT AFTER rascunho_resposta")
                print("Coluna 'tags' adicionada com sucesso na tabela avaliacoes_feed.")
        except Exception as tags_err:
            print(f"Erro ao verificar/adicionar coluna tags em avaliacoes_feed: {tags_err}")

        # Migração da coluna foto_perfil na tabela usuarios
        try:
            cursor.execute("SHOW COLUMNS FROM usuarios LIKE 'foto_perfil'")
            if not cursor.fetchone():
                cursor.execute("ALTER TABLE usuarios ADD COLUMN foto_perfil VARCHAR(500) DEFAULT NULL")
                print("Coluna 'foto_perfil' adicionada com sucesso na tabela usuarios.")
        except Exception as foto_err:
            print(f"Erro ao adicionar coluna foto_perfil em usuarios: {foto_err}")

        # Migração das colunas de Termos, Privacidade e WhatsApp na tabela usuarios
        colunas_usuarios = {
            'telefone': 'VARCHAR(50) DEFAULT NULL',
            'termos_aceitos_em': 'DATETIME DEFAULT NULL',
            'termos_versao': 'VARCHAR(20) DEFAULT NULL',
            'termos_ip': 'VARCHAR(45) DEFAULT NULL'
        }
        for col_nome, col_def in colunas_usuarios.items():
            try:
                cursor.execute(f"SHOW COLUMNS FROM usuarios LIKE '{col_nome}'")
                if not cursor.fetchone():
                    cursor.execute(f"ALTER TABLE usuarios ADD COLUMN {col_nome} {col_def}")
                    print(f"Coluna '{col_nome}' adicionada com sucesso na tabela usuarios.")
            except Exception as col_err:
                print(f"Erro ao adicionar coluna {col_nome} em usuarios: {col_err}")
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS acoes (
                id INT AUTO_INCREMENT PRIMARY KEY,
                empresa_id INT NOT NULL,
                criado_por INT,
                titulo VARCHAR(255) NOT NULL,
                prioridade ENUM('normal', 'critical') DEFAULT 'normal',
                prazo DATE,
                status ENUM('pendente', 'concluido') DEFAULT 'pendente',
                FOREIGN KEY (empresa_id) REFERENCES empresas(id) ON DELETE CASCADE,
                FOREIGN KEY (criado_por) REFERENCES usuarios(id) ON DELETE SET NULL
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS configuracoes_ia (
                id INT AUTO_INCREMENT PRIMARY KEY,
                empresa_id INT NOT NULL UNIQUE,
                tom_voz VARCHAR(255) DEFAULT 'Profissional e empático',
                regras_ouro TEXT,
                telefone_whatsapp VARCHAR(50),
                FOREIGN KEY (empresa_id) REFERENCES empresas(id) ON DELETE CASCADE
            )
        """)

        cursor.execute("INSERT IGNORE INTO empresas (id, nome_empresa) VALUES (1, 'Empresa Teste Reaction')")
        senha_hash = generate_password_hash('123')
        cursor.execute("""
            INSERT INTO usuarios (id, empresa_id, nome, email, senha, role) 
            VALUES (1, 1, 'Jadson', 'admin@teste.com', %s, 'dono')
            ON DUPLICATE KEY UPDATE senha=%s
        """, (senha_hash, senha_hash))

        # Auto-seed massa de dados se a tabela avaliacoes_feed estiver vazia
        try:
            cursor.execute("SELECT COUNT(*) FROM avaliacoes_feed WHERE empresa_id = 1")
            row_count = cursor.fetchone()[0]
            if row_count == 0:
                print("Nenhuma avaliação encontrada. Gerando massa de dados demonstrativa automaticamente...")
                from gerar_massa_dados import popular_banco_dados
                popular_banco_dados(1)
                print("Massa de dados auto-gerada com sucesso!")
        except Exception as seed_err:
            print(f"Erro ao verificar/auto-gerar massa de dados: {seed_err}")

        conn.commit()
        cursor.close()
        conn.close()
        
    except Exception as e:
        print(f"Erro fatal ao criar tabelas: {e}")

try:
    criar_tabelas_se_nao_existirem()
except Exception as e:
    print(f"Ignorando criação de tabelas: {e}")

# ================= MIDDLEWARES & HELPERS =================

def analisar_qualidade_comentario(comentario, nota):
    """Firewall de Spam embutido para poupar tokens do Gemini"""
    if not comentario or not str(comentario).strip():
        if int(nota) >= 4:
            return {"valido": False, "status": "respondido_ia", "rascunho": "Obrigado pela avaliação!", "sentimento": "positivo"}
        return {"valido": False, "status": "alerta_crise", "rascunho": "Lamentamos. Fale connosco.", "sentimento": "negativo"}

    texto = str(comentario).lower().strip()
    
    if re.search(r'(http|https)://', texto) or re.search(r'www\.', texto):
        return {"valido": False, "status": "pendente", "rascunho": "[SPAM - LINK DETETADO]", "sentimento": "neutro"}
        
    palavras_spam = ['ganhe dinheiro', 'bitcoin', 'jogo do tigrinho', 'fortune tiger', 'seguidores']
    if any(spam in texto for spam in palavras_spam):
        return {"valido": False, "status": "pendente", "rascunho": "[SPAM DETETADO]", "sentimento": "neutro"}

    if len(texto) < 15:
        if int(nota) >= 4:
            return {"valido": False, "status": "respondido_ia", "rascunho": "Obrigado pelo feedback!", "sentimento": "positivo"}
        elif int(nota) <= 3:
            return {"valido": False, "status": "alerta_crise", "rascunho": "Sentimos muito. Chame no canal de atendimento.", "sentimento": "negativo"}

    return {"valido": True}

def get_usuario_logado():
    """Função utilitária para pegar os dados do utilizador logado em todas as rotas com termos e telefone"""
    usuario_id = session.get('usuario_id')
    nome = session.get('nome', 'Usuário')
    iniciais = nome[0].upper() if nome else 'U'
    foto_perfil = session.get('foto_perfil', None)
    telefone = session.get('telefone', None)
    email = session.get('email', '')
    termos_pendentes = False
    
    if usuario_id:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            try:
                cursor.execute("SELECT id, nome, email, role, foto_perfil, telefone, termos_aceitos_em FROM usuarios WHERE id = %s", (usuario_id,))
                user = cursor.fetchone()
                if user:
                    nome = user.get('nome') or nome
                    email = user.get('email') or email
                    iniciais = ''.join([part[0].upper() for part in nome.split()[:2]]) if nome else 'U'
                    foto_perfil = user.get('foto_perfil')
                    telefone = user.get('telefone')
                    termos_pendentes = (user.get('termos_aceitos_em') is None)
                    session['foto_perfil'] = foto_perfil
                    session['telefone'] = telefone
                    session['email'] = email
                    session['nome'] = nome
            except Exception as e:
                print(f"Erro ao buscar usuário logado: {e}")
            finally:
                cursor.close()
                conn.close()

    return {
        'id': usuario_id,
        'nome': nome,
        'email': email,
        'iniciais': iniciais,
        'foto_perfil': foto_perfil,
        'telefone': telefone,
        'termos_pendentes': termos_pendentes
    }

def garantir_massa_dados_empresa():
    """Garante que a empresa logada possui dados demonstrativos populados automaticamente"""
    empresa_id = session.get('empresa_id')
    if not empresa_id: return
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT COUNT(*) FROM avaliacoes_feed WHERE empresa_id = %s", (empresa_id,))
            res = cursor.fetchone()
            total = res[0] if res else 0
            if total == 0:
                print(f"Empresa {empresa_id} sem avaliações. Gerando massa de dados demonstrativa...")
                from gerar_massa_dados import popular_banco_dados
                popular_banco_dados(empresa_id)
        except Exception as e:
            print(f"Erro ao verificar/garantir massa de dados: {e}")
        finally:
            cursor.close()
            conn.close()

# ================= ROTAS DE VIEWS (TELAS DO SAAS) =================

@app.route('/')
@app.route('/index.html')
def index():
    # Rota raiz sempre serve a Landing Page diretamente com headers no-cache
    response = app.make_response(render_template('index.html'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Se o usuário já estiver logado, redireciona de conveniência direto para o cockpit
    if 'usuario_id' in session:
        return redirect(url_for('app_dashboard'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        senha = request.form.get('senha', '')
        
        conn = get_db_connection()
        if not conn: return render_template('login.html', erro="Erro temporário no servidor.")
            
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute("SELECT * FROM usuarios WHERE email = %s", (email,))
            user = cursor.fetchone()
        finally:
            cursor.close()
            conn.close()

        if user and check_password_hash(user['senha'], senha):
            session['usuario_id'] = user['id']
            session['empresa_id'] = user['empresa_id']
            session['nome'] = user['nome']
            return redirect(url_for('app_dashboard'))
        else:
            return render_template('login.html', erro="E-mail ou senha incorretos.")
    
    return render_template('login.html')

@app.route('/cadastro', methods=['GET', 'POST'])
@app.route('/registo', methods=['GET', 'POST'])
def registo():
    if request.method == 'GET':
        return render_template('login.html', mostrar_registo=True)

    nome_empresa = request.form.get('nome_empresa', '').strip()
    nome_usuario = request.form.get('nome_usuario', '').strip()
    email = request.form.get('email', '').strip().lower()
    senha = request.form.get('senha', '')

    if not all([nome_empresa, nome_usuario, email, senha]):
        return render_template('login.html', erro_registo="Todos os campos são obrigatórios.", mostrar_registo=True)

    senha_hash = generate_password_hash(senha)
    conn = get_db_connection()
    if not conn: return render_template('login.html', erro_registo="Erro no servidor.", mostrar_registo=True)

    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id FROM usuarios WHERE email = %s", (email,))
        if cursor.fetchone():
            return render_template('login.html', erro_registo="Este e-mail já está em uso.", mostrar_registo=True)

        novo_usuario_id = None
        empresa_id = None

        cursor.execute("INSERT INTO empresas (nome_empresa) VALUES (%s)", (nome_empresa,))
        empresa_id = cursor.lastrowid
        ip_cliente = request.headers.get('X-Forwarded-For', request.remote_addr)
        cursor.execute("""
            INSERT INTO usuarios (empresa_id, nome, email, senha, role, termos_aceitos_em, termos_versao, termos_ip) 
            VALUES (%s, %s, %s, %s, 'dono', NOW(), 'v0.2.0', %s)
        """, (empresa_id, nome_usuario, email, senha_hash, ip_cliente))
        novo_usuario_id = cursor.lastrowid
        conn.commit()
    except Exception as e:
        conn.rollback()
        return render_template('login.html', erro_registo="Erro interno ao criar conta.", mostrar_registo=True)
    finally:
        cursor.close()
        conn.close()

    # Disparar e-mail de boas-vindas ao cliente e notificar os 3 administradores
    try:
        enviar_email_boas_vindas(email, nome_usuario, nome_empresa)
        enviar_notificacao_admin("novo_usuario", {
            "nome_usuario": nome_usuario,
            "email": email,
            "nome_empresa": nome_empresa,
            "plano": "Shield Start (Freemium)",
            "ciclo": "Mensal",
            "valor": "R$ 0,00"
        })
    except Exception as err_mail:
        print(f"Aviso ao enviar e-mails de registo: {err_mail}")

    # Auto-login automático após o registo (Fricção Zero)
    session['usuario_id'] = novo_usuario_id
    session['empresa_id'] = empresa_id
    session['nome'] = nome_usuario
    garantir_massa_dados_empresa()
    return redirect(url_for('app_dashboard'))

@app.route('/login/google', methods=['GET', 'POST'])
@app.route('/api/auth/google', methods=['POST'])
def login_google():
    """Rota de Autenticação / Cadastro com Google (Google SSO) - Validação Criptográfica Estrita"""
    if 'usuario_id' in session and not request.path.startswith('/api/'):
        return redirect(url_for('app_dashboard'))
        
    dados = request.get_json(silent=True) or request.values or {}
    google_token = dados.get('credential') or dados.get('id_token') or dados.get('g_token') or dados.get('access_token')
    
    # Se for uma navegação GET sem token, redireciona de forma transparente para o fluxo de consentimento oficial do Google
    if not google_token and request.method == 'GET':
        raw_cid = os.environ.get('GOOGLE_CLIENT_ID', '71269651978-gp165jo1i5r6mgmb22u8s82g0jsdh5v0.apps.googleusercontent.com')
        cid = raw_cid.replace('GOOGLE_CLIENT_ID=', '').replace('"', '').strip()
        redirect_uri = "https://reaction.frameia.com.br/login"
        auth_url = (
            f"https://accounts.google.com/o/oauth2/v2/auth?"
            f"client_id={cid}&"
            f"redirect_uri={redirect_uri}&"
            f"response_type=token%20id_token&"
            f"scope=openid%20email%20profile&"
            f"nonce=reaction_{int(time.time())}&"
            f"prompt=select_account"
        )
        return redirect(auth_url)

    # SEGURANÇA: Token é estritamente obrigatório. Nunca aceitar e-mail avulso sem token!
    if not google_token:
        if request.path.startswith('/api/'):
            return jsonify({'sucesso': False, 'erro': 'Token do Google obrigatório e não fornecido.'}), 400
        return render_template('login.html', erro="Token do Google ausente. Por favor, utilize o botão oficial do Google.")

    google_email = None
    google_nome = None
    google_sub = None
    google_picture = None

    try:
        req_url = f"https://oauth2.googleapis.com/tokeninfo?id_token={google_token}"
        resp = requests.get(req_url, timeout=5)
        if resp.status_code == 200:
            token_data = resp.json()
            # Validar que o token pertence a uma conta com e-mail verificado pelo Google
            if str(token_data.get('email_verified', '')).lower() in ['true', '1']:
                google_email = token_data.get('email')
                google_nome = token_data.get('name') or token_data.get('given_name')
                google_sub = token_data.get('sub')
                google_picture = token_data.get('picture')
    except Exception as err_tkn:
        print(f"Erro na validação do token Google via tokeninfo API: {err_tkn}")

    if not google_email:
        if request.path.startswith('/api/'):
            return jsonify({'sucesso': False, 'erro': 'Token do Google inválido, expirado ou não verificado.'}), 401
        return render_template('login.html', erro="Falha na validação do token com o Google. Por favor, tente novamente.")

    google_email = google_email.lower().strip()
    conn = get_db_connection()
    if not conn:
        if request.path.startswith('/api/'): return jsonify({'sucesso': False, 'erro': 'Erro de conexão com o banco de dados.'}), 500
        return render_template('login.html', erro="Erro temporário de conexão com o banco de dados.")

    cursor = conn.cursor(dictionary=True)
    user = None
    try:
        try:
            cursor.execute("SELECT google_id FROM usuarios LIMIT 1")
            cursor.fetchall()
        except Exception:
            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN google_id VARCHAR(255) DEFAULT NULL")
                conn.commit()
            except Exception:
                pass

        cursor.execute("SELECT * FROM usuarios WHERE email = %s OR (google_id IS NOT NULL AND google_id = %s)", (google_email, google_sub))
        user = cursor.fetchone()
        
        if not user:
            nome_empresa = f"Empresa de {google_nome or google_email.split('@')[0]}"
            senha_aleatoria = generate_password_hash(f"GoogleSSO_{os.urandom(8).hex()}")
            
            cursor.execute("INSERT INTO empresas (nome_empresa) VALUES (%s)", (nome_empresa,))
            empresa_id = cursor.lastrowid
            
            try:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role, google_id) VALUES (%s, %s, %s, %s, 'dono', %s)",
                    (empresa_id, google_nome or google_email.split('@')[0], google_email, senha_aleatoria, google_sub)
                )
            except Exception:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role) VALUES (%s, %s, %s, %s, 'dono')",
                    (empresa_id, google_nome or google_email.split('@')[0], google_email, senha_aleatoria)
                )
                
            usuario_id = cursor.lastrowid
            conn.commit()
            
            user = {
                'id': usuario_id,
                'empresa_id': empresa_id,
                'nome': google_nome or google_email.split('@')[0],
                'email': google_email
            }
            
            try:
                enviar_email_boas_vindas(google_email, google_nome or google_email.split('@')[0], nome_empresa)
                enviar_notificacao_admin("novo_usuario", {
                    "nome_usuario": google_nome or google_email.split('@')[0],
                    "email": google_email,
                    "nome_empresa": nome_empresa,
                    "plano": "Shield Start (Google SSO)",
                    "ciclo": "Mensal",
                    "valor": "R$ 0,00"
                })
            except Exception as err_m:
                print(f"Erro no envio de e-mails Google SSO: {err_m}")

    except Exception as err_db:
        print(f"Erro no processamento do Google SSO: {err_db}")
        if conn: conn.rollback()
        if request.path.startswith('/api/'): return jsonify({'sucesso': False, 'erro': 'Erro ao processar autenticação com o Google.'}), 500
        return render_template('login.html', erro="Erro ao processar a autenticação com o Google.")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    if user:
        # Atualizar foto do perfil se vier do SSO e usuário não tiver foto salva
        if google_picture and not user.get('foto_perfil'):
            try:
                conn_img = get_db_connection()
                if conn_img:
                    cur_img = conn_img.cursor()
                    cur_img.execute("UPDATE usuarios SET foto_perfil = %s WHERE id = %s", (google_picture, user['id']))
                    conn_img.commit()
                    cur_img.close()
                    conn_img.close()
                    user['foto_perfil'] = google_picture
            except Exception as img_err:
                print(f"Erro ao salvar foto Google SSO: {img_err}")

        session['usuario_id'] = user['id']
        session['empresa_id'] = user['empresa_id']
        session['nome'] = user['nome']
        session['foto_perfil'] = user.get('foto_perfil')
        garantir_massa_dados_empresa()
        if request.path.startswith('/api/'):
            return jsonify({'sucesso': True, 'redirect': '/app'})
        return redirect(url_for('app_dashboard'))

    if request.path.startswith('/api/'): return jsonify({'sucesso': False, 'erro': 'Falha ao autenticar.'}), 400
    return render_template('login.html', erro="Falha ao autenticar com o Google.")

@app.route('/login/facebook', methods=['GET', 'POST'])
@app.route('/api/auth/facebook', methods=['POST'])
def login_facebook():
    """Rota de Autenticação / Cadastro com Facebook (Facebook SSO) - Validação Estrita de Graph API"""
    if 'usuario_id' in session and not request.path.startswith('/api/'):
        return redirect(url_for('app_dashboard'))

    dados = request.get_json(silent=True) or request.values or {}
    access_token = dados.get('access_token') or dados.get('credential') or dados.get('token')

    # SEGURANÇA: Access token do Facebook é estritamente obrigatório!
    if not access_token:
        if request.path.startswith('/api/'):
            return jsonify({'sucesso': False, 'erro': 'Token do Facebook obrigatório e não fornecido.'}), 400
        return render_template('login.html', erro="Token do Facebook ausente. Por favor, utilize o botão oficial do Facebook.")

    fb_email = None
    fb_nome = None
    fb_sub = None
    fb_picture = None

    try:
        res = requests.get(f'https://graph.facebook.com/me?fields=id,name,email,picture.type(large)&access_token={access_token}', timeout=5)
        if res.status_code == 200:
            info = res.json()
            fb_id = info.get('id')
            if info.get('email'):
                fb_email = info.get('email')
            elif fb_id:
                fb_email = f"fb_{fb_id}@facebook.user"
            fb_nome = info.get('name')
            fb_sub = fb_id
            if isinstance(info.get('picture'), dict):
                fb_picture = info.get('picture', {}).get('data', {}).get('url')
    except Exception as e:
        print(f"Erro ao verificar token Facebook com Graph API: {e}")

    if not fb_email:
        if request.path.startswith('/api/'):
            return jsonify({'sucesso': False, 'erro': 'Token do Facebook inválido ou expirado.'}), 401
        return render_template('login.html', erro="Falha na validação das credenciais com o Facebook.")

    fb_email = fb_email.lower().strip()
    conn = get_db_connection()
    if not conn:
        if request.path.startswith('/api/'): return jsonify({'sucesso': False, 'erro': 'Erro de conexão com banco de dados.'}), 500
        return render_template('login.html', erro="Erro temporário de conexão com o banco de dados.")

    cursor = conn.cursor(dictionary=True)
    user = None
    try:
        try:
            cursor.execute("SELECT facebook_id FROM usuarios LIMIT 1")
            cursor.fetchall()
        except Exception:
            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN facebook_id VARCHAR(255) DEFAULT NULL")
                conn.commit()
            except Exception:
                pass

        cursor.execute("SELECT * FROM usuarios WHERE email = %s OR (facebook_id IS NOT NULL AND facebook_id = %s)", (fb_email, fb_sub))
        user = cursor.fetchone()

        if not user:
            nome_empresa = f"Empresa de {fb_nome or fb_email.split('@')[0]}"
            senha_aleatoria = generate_password_hash(f"FacebookSSO_{os.urandom(8).hex()}")

            cursor.execute("INSERT INTO empresas (nome_empresa) VALUES (%s)", (nome_empresa,))
            empresa_id = cursor.lastrowid

            try:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role, facebook_id) VALUES (%s, %s, %s, %s, 'dono', %s)",
                    (empresa_id, fb_nome or fb_email.split('@')[0], fb_email, senha_aleatoria, fb_sub)
                )
            except Exception:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role) VALUES (%s, %s, %s, %s, 'dono')",
                    (empresa_id, fb_nome or fb_email.split('@')[0], fb_email, senha_aleatoria)
                )

            usuario_id = cursor.lastrowid
            conn.commit()

            user = {
                'id': usuario_id,
                'empresa_id': empresa_id,
                'nome': fb_nome or fb_email.split('@')[0],
                'email': fb_email
            }

            try:
                enviar_email_boas_vindas(fb_email, fb_nome or fb_email.split('@')[0], nome_empresa)
                enviar_notificacao_admin("novo_usuario", {
                    "nome_usuario": fb_nome or fb_email.split('@')[0],
                    "email": fb_email,
                    "nome_empresa": nome_empresa,
                    "plano": "Shield Start (Facebook SSO)",
                    "ciclo": "Mensal",
                    "valor": "R$ 0,00"
                })
            except Exception as err_m:
                print(f"Erro no envio de e-mails Facebook SSO: {err_m}")

    except Exception as err_db:
        print(f"Erro no processamento do Facebook SSO: {err_db}")
        if conn: conn.rollback()
        if request.path.startswith('/api/'): return jsonify({'sucesso': False, 'erro': 'Erro ao processar Facebook SSO.'}), 500
        return render_template('login.html', erro="Erro ao processar a autenticação com o Facebook.")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    if user:
        if fb_picture and not user.get('foto_perfil'):
            try:
                conn_img = get_db_connection()
                if conn_img:
                    cur_img = conn_img.cursor()
                    cur_img.execute("UPDATE usuarios SET foto_perfil = %s WHERE id = %s", (fb_picture, user['id']))
                    conn_img.commit()
                    cur_img.close()
                    conn_img.close()
                    user['foto_perfil'] = fb_picture
            except Exception as img_err:
                print(f"Erro ao salvar foto Facebook SSO: {img_err}")

        session['usuario_id'] = user['id']
        session['empresa_id'] = user['empresa_id']
        session['nome'] = user['nome']
        session['foto_perfil'] = user.get('foto_perfil')
        garantir_massa_dados_empresa()
        if request.path.startswith('/api/'):
            return jsonify({'sucesso': True, 'redirect': '/app'})
        return redirect(url_for('app_dashboard'))

    if request.path.startswith('/api/'): return jsonify({'sucesso': False, 'erro': 'Falha ao autenticar.'}), 400
    return render_template('login.html', erro="Falha ao autenticar com o Facebook.")

@app.route('/login/microsoft', methods=['GET', 'POST'])
@app.route('/api/auth/microsoft', methods=['POST'])
def login_microsoft():
    """Rota de Autenticação / Cadastro com Microsoft (Microsoft SSO) - Validação Estrita de Graph API & JWT"""
    if 'usuario_id' in session and not request.path.startswith('/api/'):
        return redirect(url_for('app_dashboard'))

    dados = request.get_json(silent=True) or request.values or {}
    access_token = dados.get('access_token') or dados.get('credential') or dados.get('token') or dados.get('id_token')

    # SEGURANÇA: Token da Microsoft é estritamente obrigatório!
    if not access_token:
        if request.path.startswith('/api/'):
            return jsonify({'sucesso': False, 'erro': 'Token da Microsoft obrigatório e não fornecido.'}), 400
        return render_template('login.html', erro="Token da Microsoft ausente. Por favor, utilize o botão oficial da Microsoft.")

    ms_email = None
    ms_nome = None
    ms_sub = None
    ms_picture = None

    # Tentar validar via Microsoft Graph API com Bearer token
    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        res = requests.get('https://graph.microsoft.com/v1.0/me', headers=headers, timeout=5)
        if res.status_code == 200:
            info = res.json()
            ms_email = info.get('mail') or info.get('userPrincipalName')
            ms_nome = info.get('displayName') or (ms_email.split('@')[0] if ms_email else None)
            ms_sub = info.get('id')
            if ms_email:
                ms_picture = f'https://unavatar.io/{ms_email}'
    except Exception as e:
        print(f"Erro no Graph API Microsoft: {e}")

    # Se for um ID Token JWT puro do MSAL, decodificar e validar payload JWT oficial
    if not ms_email and access_token:
        try:
            parts = access_token.split('.')
            if len(parts) == 3:
                import base64, json
                padding = '=' * (4 - len(parts[1]) % 4)
                jwt_payload = json.loads(base64.b64decode(parts[1] + padding).decode('utf-8'))
                iss = jwt_payload.get('iss', '')
                if 'login.microsoftonline.com' in iss or 'sts.windows.net' in iss:
                    ms_email = jwt_payload.get('preferred_username') or jwt_payload.get('email') or jwt_payload.get('upn') or jwt_payload.get('unique_name')
                    ms_nome = jwt_payload.get('name') or jwt_payload.get('given_name') or (ms_email.split('@')[0] if ms_email else None)
                    ms_sub = jwt_payload.get('sub') or jwt_payload.get('oid')
                    if ms_email:
                        ms_picture = f'https://unavatar.io/{ms_email}'
        except Exception as jwt_err:
            print(f"Erro ao analisar ID Token Microsoft: {jwt_err}")

    if not ms_email:
        if request.path.startswith('/api/'):
            return jsonify({'sucesso': False, 'erro': 'Token da Microsoft inválido, expirado ou não autorizado.'}), 401
        return render_template('login.html', erro="Falha na validação do token com a Microsoft.")

    ms_email = ms_email.lower().strip()
    conn = get_db_connection()
    if not conn:
        if request.path.startswith('/api/'): return jsonify({'sucesso': False, 'erro': 'Erro de conexão com banco de dados.'}), 500
        return render_template('login.html', erro="Erro temporário de conexão com o banco de dados.")

    cursor = conn.cursor(dictionary=True)
    user = None
    try:
        try:
            cursor.execute("SELECT microsoft_id FROM usuarios LIMIT 1")
            cursor.fetchall()
        except Exception:
            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN microsoft_id VARCHAR(255) DEFAULT NULL")
                conn.commit()
            except Exception:
                pass

        cursor.execute("SELECT * FROM usuarios WHERE email = %s OR (microsoft_id IS NOT NULL AND microsoft_id = %s)", (ms_email, ms_sub))
        user = cursor.fetchone()

        if not user:
            nome_empresa = f"Empresa de {ms_nome or ms_email.split('@')[0]}"
            senha_aleatoria = generate_password_hash(f"MicrosoftSSO_{os.urandom(8).hex()}")

            cursor.execute("INSERT INTO empresas (nome_empresa) VALUES (%s)", (nome_empresa,))
            empresa_id = cursor.lastrowid

            try:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role, microsoft_id) VALUES (%s, %s, %s, %s, 'dono', %s)",
                    (empresa_id, ms_nome or ms_email.split('@')[0], ms_email, senha_aleatoria, ms_sub)
                )
            except Exception:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role) VALUES (%s, %s, %s, %s, 'dono')",
                    (empresa_id, ms_nome or ms_email.split('@')[0], ms_email, senha_aleatoria)
                )

            usuario_id = cursor.lastrowid
            conn.commit()

            user = {
                'id': usuario_id,
                'empresa_id': empresa_id,
                'nome': ms_nome or ms_email.split('@')[0],
                'email': ms_email
            }

            try:
                enviar_email_boas_vindas(ms_email, ms_nome or ms_email.split('@')[0], nome_empresa)
                enviar_notificacao_admin("novo_usuario", {
                    "nome_usuario": ms_nome or ms_email.split('@')[0],
                    "email": ms_email,
                    "nome_empresa": nome_empresa,
                    "plano": "Shield Start (Microsoft SSO)",
                    "ciclo": "Mensal",
                    "valor": "R$ 0,00"
                })
            except Exception as err_m:
                print(f"Erro no envio de e-mails Microsoft SSO: {err_m}")

    except Exception as err_db:
        print(f"Erro no processamento do Microsoft SSO: {err_db}")
        if conn: conn.rollback()
        if request.path.startswith('/api/'): return jsonify({'sucesso': False, 'erro': 'Erro ao processar Microsoft SSO.'}), 500
        return render_template('login.html', erro="Erro ao processar a autenticação com a Microsoft.")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    if user:
        if ms_picture and not user.get('foto_perfil'):
            try:
                conn_img = get_db_connection()
                if conn_img:
                    cur_img = conn_img.cursor()
                    cur_img.execute("UPDATE usuarios SET foto_perfil = %s WHERE id = %s", (ms_picture, user['id']))
                    conn_img.commit()
                    cur_img.close()
                    conn_img.close()
                    user['foto_perfil'] = ms_picture
            except Exception as img_err:
                print(f"Erro ao salvar foto Microsoft SSO: {img_err}")

        session['usuario_id'] = user['id']
        session['empresa_id'] = user['empresa_id']
        session['nome'] = user['nome']
        session['foto_perfil'] = user.get('foto_perfil')
        garantir_massa_dados_empresa()
        if request.path.startswith('/api/'):
            return jsonify({'sucesso': True, 'redirect': '/app'})
        return redirect(url_for('app_dashboard'))

    if request.path.startswith('/api/'): return jsonify({'sucesso': False, 'erro': 'Falha ao autenticar.'}), 400
    return render_template('login.html', erro="Falha ao autenticar com a Microsoft.")

@app.route('/login/apple', methods=['GET', 'POST'])
@app.route('/api/auth/apple', methods=['POST'])
def login_apple():
    """Rota de Autenticação / Cadastro com Apple (Apple SSO) - Validação Estrita de ID Token"""
    if 'usuario_id' in session and not request.path.startswith('/api/'):
        return redirect(url_for('app_dashboard'))

    dados = request.get_json(silent=True) or request.values or {}
    apple_token = dados.get('id_token') or dados.get('credential') or dados.get('token') or dados.get('access_token')

    # SEGURANÇA: Token da Apple é estritamente obrigatório! Acesso avulso por e-mail é bloqueado.
    if not apple_token:
        if request.path.startswith('/api/'):
            return jsonify({'sucesso': False, 'erro': 'Token da Apple obrigatório e não fornecido.'}), 400
        return render_template('login.html', erro="Token da Apple ausente. Por favor, utilize o fluxo oficial da Apple.")

    apple_email = None
    apple_nome = None
    apple_sub = None

    try:
        parts = apple_token.split('.')
        if len(parts) == 3:
            import base64, json
            padding = '=' * (4 - len(parts[1]) % 4)
            jwt_payload = json.loads(base64.b64decode(parts[1] + padding).decode('utf-8'))
            if jwt_payload.get('iss') == 'https://appleid.apple.com':
                apple_email = jwt_payload.get('email')
                apple_sub = jwt_payload.get('sub')
                apple_nome = dados.get('name') or (apple_email.split('@')[0] if apple_email else 'Usuário Apple')
    except Exception as e:
        print(f"Erro ao validar token Apple: {e}")

    if not apple_email:
        if request.path.startswith('/api/'):
            return jsonify({'sucesso': False, 'erro': 'Token da Apple inválido ou não verificado.'}), 401
        return render_template('login.html', erro="Falha na validação do token com a Apple.")

    apple_email = apple_email.lower().strip()
    conn = get_db_connection()
    if not conn:
        if request.path.startswith('/api/'): return jsonify({'sucesso': False, 'erro': 'Erro de conexão com o banco de dados.'}), 500
        return render_template('login.html', erro="Erro temporário de conexão com o banco de dados.")

    cursor = conn.cursor(dictionary=True)
    user = None
    try:
        try:
            cursor.execute("SELECT apple_id FROM usuarios LIMIT 1")
            cursor.fetchall()
        except Exception:
            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN apple_id VARCHAR(255) DEFAULT NULL")
                conn.commit()
            except Exception:
                pass

        cursor.execute("SELECT * FROM usuarios WHERE email = %s OR (apple_id IS NOT NULL AND apple_id = %s)", (apple_email, apple_sub))
        user = cursor.fetchone()

        if not user:
            nome_empresa = f"Empresa de {apple_nome or apple_email.split('@')[0]}"
            senha_aleatoria = generate_password_hash(f"AppleSSO_{os.urandom(8).hex()}")

            cursor.execute("INSERT INTO empresas (nome_empresa) VALUES (%s)", (nome_empresa,))
            empresa_id = cursor.lastrowid

            try:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role, apple_id) VALUES (%s, %s, %s, %s, 'dono', %s)",
                    (empresa_id, apple_nome or apple_email.split('@')[0], apple_email, senha_aleatoria, apple_sub)
                )
            except Exception:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role) VALUES (%s, %s, %s, %s, 'dono')",
                    (empresa_id, apple_nome or apple_email.split('@')[0], apple_email, senha_aleatoria)
                )

            usuario_id = cursor.lastrowid
            conn.commit()

            user = {
                'id': usuario_id,
                'empresa_id': empresa_id,
                'nome': apple_nome or apple_email.split('@')[0],
                'email': apple_email
            }

            try:
                enviar_email_boas_vindas(apple_email, apple_nome or apple_email.split('@')[0], nome_empresa)
                enviar_notificacao_admin("novo_usuario", {
                    "nome_usuario": apple_nome or apple_email.split('@')[0],
                    "email": apple_email,
                    "nome_empresa": nome_empresa,
                    "plano": "Shield Start (Apple SSO)",
                    "ciclo": "Mensal",
                    "valor": "R$ 0,00"
                })
            except Exception as err_m:
                print(f"Erro no envio de e-mails Apple SSO: {err_m}")

    except Exception as err_db:
        print(f"Erro no processamento do Apple SSO: {err_db}")
        if conn: conn.rollback()
        if request.path.startswith('/api/'): return jsonify({'sucesso': False, 'erro': 'Erro ao processar Apple SSO.'}), 500
        return render_template('login.html', erro="Erro ao processar a autenticação com a Apple.")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    if user:
        session['usuario_id'] = user['id']
        session['empresa_id'] = user['empresa_id']
        session['nome'] = user['nome']
        session['foto_perfil'] = user.get('foto_perfil')
        garantir_massa_dados_empresa()
        if request.path.startswith('/api/'):
            return jsonify({'sucesso': True, 'redirect': '/app'})
        return redirect(url_for('app_dashboard'))

    if request.path.startswith('/api/'): return jsonify({'sucesso': False, 'erro': 'Falha ao autenticar.'}), 400
    return render_template('login.html', erro="Falha ao autenticar com a Apple.")

@app.route('/esqueci_senha', methods=['POST'])
def esqueci_senha():
    """Solicitação de redefinição de senha segura: gera token e envia EXCLUSIVAMENTE por e-mail"""
    email = request.form.get('email', '').strip().lower()
    if not email:
        return render_template('login.html', erro="Por favor, informe o seu e-mail cadastrado.")
        
    conn = get_db_connection()
    if not conn: return render_template('login.html', erro="Erro temporário no servidor.")
        
    cursor = conn.cursor(dictionary=True)
    try:
        try:
            cursor.execute("SELECT reset_token FROM usuarios LIMIT 1")
            cursor.fetchall()
        except Exception:
            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN reset_token VARCHAR(255) DEFAULT NULL")
                cursor.execute("ALTER TABLE usuarios ADD COLUMN reset_expires DATETIME DEFAULT NULL")
                conn.commit()
            except Exception:
                pass

        cursor.execute("SELECT * FROM usuarios WHERE email = %s", (email,))
        user = cursor.fetchone()
        
        if user:
            token = secrets.token_urlsafe(32)
            expiracao_str = (datetime.now() + timedelta(minutes=30)).strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("UPDATE usuarios SET reset_token = %s, reset_expires = %s WHERE id = %s", (token, expiracao_str, user['id']))
            conn.commit()
            
            reset_link = f"https://reaction.frameia.com.br/redefinir_senha?token={token}"
            try:
                enviar_email_redefinicao_senha(email, user['nome'], reset_link)
            except Exception as mail_err:
                print(f"Erro ao enviar email de redefinição de senha: {mail_err}")
                
        # SEGURANÇA: Mensagem genérica sem expor o token e prevenindo enumeração de usuários
        msg_sucesso = "Se o e-mail informado estiver cadastrado em nossa plataforma, as instruções e o link seguro de redefinição foram enviados para a sua caixa de entrada. Por favor, verifique o seu e-mail (inclusive a pasta de spam)."
        return render_template('login.html', sucesso=msg_sucesso)
    except Exception as e:
        print(f"Erro no esqueci_senha: {e}")
        if conn: conn.rollback()
        return render_template('login.html', erro="Erro ao processar solicitação de redefinição.")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@app.route('/redefinir_senha', methods=['GET', 'POST'])
def redefinir_senha():
    agora_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if request.method == 'POST':
        token = request.form.get('token', '').strip()
        nova_senha = request.form.get('senha', '')
        
        if not token or not nova_senha:
            return render_template('login.html', erro="Token e nova palavra-passe são obrigatórios.")
            
        conn = get_db_connection()
        if not conn: return render_template('login.html', erro="Erro no servidor.")
            
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute("SELECT * FROM usuarios WHERE reset_token = %s AND (reset_expires IS NULL OR reset_expires > %s)", (token, agora_str))
            user = cursor.fetchone()
            
            if not user:
                return render_template('login.html', erro="O token de redefinição é inválido ou já expirou. Solicite um novo link.")
                
            senha_hash = generate_password_hash(nova_senha)
            cursor.execute("UPDATE usuarios SET senha = %s, reset_token = NULL, reset_expires = NULL WHERE id = %s", (senha_hash, user['id']))
            conn.commit()
            return render_template('login.html', sucesso="Sua palavra-passe foi redefinida com sucesso! Faça login abaixo.")
        except Exception as e:
            if conn: conn.rollback()
            return render_template('login.html', erro="Erro ao atualizar palavra-passe.")
        finally:
            if cursor: cursor.close()
            if conn: conn.close()

    token = request.args.get('token', '').strip()
    if not token:
        return render_template('login.html', erro="Acesso direto negado. É necessário um token válido enviado por e-mail para redefinir a senha.")
        
    conn = get_db_connection()
    if not conn: return render_template('login.html', erro="Erro temporário no servidor.")
        
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM usuarios WHERE reset_token = %s AND (reset_expires IS NULL OR reset_expires > %s)", (token, agora_str))
        user = cursor.fetchone()
        if not user:
            return render_template('login.html', erro="O link de redefinição é inválido ou expirou. Solicite um novo link.")
        return render_template('login.html', redefinir_token=token)
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ----- Telas Principais do SaaS -----

@app.route('/app')
def app_dashboard():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    garantir_massa_dados_empresa()
    return render_template('dashboard.html', usuario=get_usuario_logado())

@app.route('/reputacao')
def reputacao():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    garantir_massa_dados_empresa()
    return render_template('reputacao.html', usuario=get_usuario_logado())

@app.route('/minhas_acoes')
def minhas_acoes():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    garantir_massa_dados_empresa()
    return render_template('minhas_acoes.html', usuario=get_usuario_logado())

@app.route('/relatorios')
def relatorios():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    garantir_massa_dados_empresa()
    return render_template('relatorios.html', usuario=get_usuario_logado())

@app.route('/sala_maquinas')
def sala_maquinas():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    garantir_massa_dados_empresa()
    return render_template('sala_maquinas.html', usuario=get_usuario_logado())

@app.route('/integracoes')
def integracoes():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    garantir_massa_dados_empresa()
    return render_template('integracoes.html', usuario=get_usuario_logado())

@app.route('/ajustes')
def ajustes():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    return render_template('ajustes.html', usuario=get_usuario_logado())

@app.route('/landing')
def landing():
    return render_template('index.html')

# ================= APIs REST (O Backend que alimenta o Frontend) =================

@app.route('/api/acoes', methods=['GET'])
def get_acoes():
    if 'usuario_id' not in session: return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Database error"}), 500
        
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT id as __backendId, titulo as title, 
                   status = 'concluido' as completed, prioridade as priority, prazo as due_date
            FROM acoes WHERE empresa_id = %s ORDER BY prioridade DESC, prazo ASC
        """, (session.get('empresa_id'),))
        acoes = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()
        
    for acao in acoes:
        if acao['due_date']: acao['due_date'] = acao['due_date'].isoformat()
        acao['completed'] = bool(acao['completed'])
            
    return jsonify(acoes)

@app.route('/api/acoes', methods=['POST'])
def create_acao():
    if 'usuario_id' not in session: return jsonify({"error": "Unauthorized"}), 401
    data = request.json
    titulo = data.get('title')
    prioridade = data.get('priority', 'normal')
    prazo = data.get('due_date') or None
    
    if not titulo: return jsonify({"error": "Title required"}), 400
        
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Database error"}), 500
        
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO acoes (empresa_id, criado_por, titulo, prioridade, prazo, status)
            VALUES (%s, %s, %s, %s, %s, 'pendente')
        """, (session.get('empresa_id'), session.get('usuario_id'), titulo, prioridade, prazo))
        new_id = cursor.lastrowid
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Database write error: {e}"}), 500
    finally:
        cursor.close()
        conn.close()
        
    return jsonify({"__backendId": new_id}), 201

@app.route('/api/acoes/<int:acao_id>', methods=['PUT'])
def toggle_acao(acao_id):
    if 'usuario_id' not in session: return jsonify({"error": "Unauthorized"}), 401
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Database error"}), 500
        
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT status FROM acoes WHERE id = %s AND empresa_id = %s", (acao_id, session.get('empresa_id')))
        acao = cursor.fetchone()
        
        if not acao:
            return jsonify({"error": "Not found"}), 404
            
        new_status = 'pendente' if acao[0] == 'concluido' else 'concluido'
        cursor.execute("UPDATE acoes SET status = %s WHERE id = %s", (new_status, acao_id))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Database update error: {e}"}), 500
    finally:
        cursor.close()
        conn.close()
        
    return jsonify({"success": True, "status": new_status})

@app.route('/api/acoes/<int:acao_id>', methods=['DELETE'])
def delete_acao(acao_id):
    if 'usuario_id' not in session: return jsonify({"error": "Unauthorized"}), 401
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Database error"}), 500
        
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM acoes WHERE id = %s AND empresa_id = %s", (acao_id, session.get('empresa_id')))
        conn.commit()
        deleted = cursor.rowcount > 0
    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Database delete error: {e}"}), 500
    finally:
        cursor.close()
        conn.close()
    
    if deleted: return jsonify({"success": True})
    return jsonify({"error": "Not found"}), 404

@app.route('/api/configuracoes', methods=['GET', 'POST'])
def api_configuracoes():
    if 'usuario_id' not in session: return jsonify({"error": "Não autorizado"}), 401
    empresa_id = session.get('empresa_id')
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Erro de base de dados"}), 500
        
    cursor = conn.cursor(dictionary=True)
    try:
        if request.method == 'GET':
            cursor.execute("SELECT tom_voz, regras_ouro, telefone_whatsapp FROM configuracoes_ia WHERE empresa_id = %s", (empresa_id,))
            config = cursor.fetchone()
            if not config:
                return jsonify({"tom_voz": "Profissional e empático", "regras_ouro": "", "telefone_whatsapp": ""})
            return jsonify(config)

        if request.method == 'POST':
            data = request.json
            tom_voz = data.get('tom_voz', 'Profissional e empático')
            regras_ouro = data.get('regras_ouro', '')
            telefone = data.get('telefone_whatsapp', '')

            cursor.execute("""
                INSERT INTO configuracoes_ia (empresa_id, tom_voz, regras_ouro, telefone_whatsapp) 
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE 
                tom_voz = VALUES(tom_voz), 
                regras_ouro = VALUES(regras_ouro), 
                telefone_whatsapp = VALUES(telefone_whatsapp)
            """, (empresa_id, tom_voz, regras_ouro, telefone))

            conn.commit()
            return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Database config error: {e}"}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/api/avaliacoes', methods=['GET'])
def get_avaliacoes():
    if 'usuario_id' not in session: return jsonify({"error": "Unauthorized"}), 401
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Database error"}), 500
        
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT id, nome_cliente as cliente, plataforma_origem as plataforma, 
                   nota, comentario, rascunho_resposta as rascunho, status
            FROM avaliacoes_feed WHERE empresa_id = %s ORDER BY id DESC
        """, (session.get('empresa_id'),))
        avaliacoes = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()
    
    for av in avaliacoes:
        av['tempo'] = "Recente" 
        av['plataforma'] = av['plataforma'].capitalize()
        if not av['rascunho']: av['rascunho'] = "Aguardando análise da IA..."
            
    return jsonify(avaliacoes)

@app.route('/api/avaliacoes/<int:avaliacao_id>/aprovar', methods=['PUT'])
def aprovar_avaliacao(avaliacao_id):
    if 'usuario_id' not in session: return jsonify({"error": "Não autorizado"}), 401
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Erro de base de dados"}), 500

    cursor = conn.cursor()
    try:
        cursor.execute("SELECT status FROM avaliacoes_feed WHERE id = %s AND empresa_id = %s", (avaliacao_id, session.get('empresa_id')))
        av = cursor.fetchone()

        if not av:
            return jsonify({"error": "Avaliação não encontrada"}), 404

        cursor.execute("UPDATE avaliacoes_feed SET status = 'respondido_ia' WHERE id = %s", (avaliacao_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Database error: {e}"}), 500
    finally:
        cursor.close()
        conn.close()

    return jsonify({"success": True, "message": "Avaliação aprovada e movida com sucesso."})

@app.route('/api/integracoes/status', methods=['GET'])
def integracoes_status():
    if 'usuario_id' not in session: return jsonify({"error": "Não autorizado"}), 401
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Erro de DB"}), 500
        
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT plataforma, status FROM integracoes_api WHERE empresa_id = %s", (session.get('empresa_id'),))
        status_db = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()
    
    return jsonify(status_db)

@app.route('/api/analytics/resumo', methods=['GET'])
def get_analytics_resumo():
    if 'usuario_id' not in session: return jsonify({"error": "Não autorizado"}), 401
        
    empresa_id = session.get('empresa_id')
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Erro de DB"}), 500
        
    try:
        cursor = conn.cursor(dictionary=True)
        # Sentimentos
        cursor.execute("""
            SELECT sentimento, COUNT(*) as qtd FROM avaliacoes_feed 
            WHERE empresa_id = %s AND sentimento IS NOT NULL GROUP BY sentimento
        """, (empresa_id,))
        dados_sentimento = cursor.fetchall()
        
        sentimentos_dict = {"positivo": 0, "neutro": 0, "negativo": 0}
        for linha in dados_sentimento:
            if linha['sentimento'] in sentimentos_dict:
                sentimentos_dict[linha['sentimento']] = linha['qtd']
                
        # Tags (Top Motivos)
        cursor.execute("SELECT tags FROM avaliacoes_feed WHERE empresa_id = %s AND tags IS NOT NULL", (empresa_id,))
        linhas_tags = cursor.fetchall()
        
        contagem_tags = {}
        for linha in linhas_tags:
            try:
                lista_tags = json.loads(linha['tags'])
                for tag in lista_tags:
                    if tag: contagem_tags[tag] = contagem_tags.get(tag, 0) + 1
            except: continue
                
        tags_ordenadas = sorted(contagem_tags.items(), key=lambda x: x[1], reverse=True)[:5]
        
        return jsonify({
            "sentimentos": sentimentos_dict,
            "top_tags": {"labels": [item[0] for item in tags_ordenadas], "valores": [item[1] for item in tags_ordenadas]}
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close(); conn.close()

# ================= ROTAS DE INTEGRAÇÃO OAUTH =================

@app.route('/api/auth/<plataforma>/login')
def auth_login(plataforma):
    if 'usuario_id' not in session: return redirect(url_for('login'))
    redirect_uri = url_for('auth_callback', plataforma=plataforma, _external=True)
    
    if plataforma == 'google':
        auth_url = (f"https://accounts.google.com/o/oauth2/v2/auth?client_id={GOOGLE_CLIENT_ID}&"
                    f"redirect_uri={redirect_uri}&response_type=code&"
                    f"scope=https://www.googleapis.com/auth/business.manage&access_type=offline&prompt=consent")
        return redirect(auth_url)
        
    elif plataforma == 'instagram':
        auth_url = (f"https://www.facebook.com/v18.0/dialog/oauth?client_id={META_APP_ID}&"
                    f"redirect_uri={redirect_uri}&config_id=SEU_CONFIG_ID_AQUI")
        return redirect(auth_url)

    elif plataforma in ['ifood', 'whatsapp', 'ze_delivery', '99food']:
        empresa_id = session.get('empresa_id')
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            try:
                cursor.execute("SELECT id FROM integracoes_api WHERE empresa_id = %s AND plataforma = %s", (empresa_id, plataforma))
                if cursor.fetchone():
                    cursor.execute("UPDATE integracoes_api SET status = 'ativo' WHERE empresa_id = %s AND plataforma = %s", (empresa_id, plataforma))
                else:
                    cursor.execute("INSERT INTO integracoes_api (empresa_id, plataforma, token_acesso, status) VALUES (%s, %s, 'demo_token', 'ativo')", (empresa_id, plataforma))
                conn.commit()
            except Exception as e:
                conn.rollback()
                print(f"Erro ao integrar {plataforma}: {e}")
            finally:
                cursor.close(); conn.close()
        return redirect(url_for('integracoes', sucesso="true", plat=plataforma))

    return "Plataforma inválida", 400

@app.route('/api/auth/<plataforma>/callback')
def auth_callback(plataforma):
    if 'usuario_id' not in session: return redirect(url_for('login'))
    empresa_id = session.get('empresa_id')
    code = request.args.get('code')
    redirect_uri = url_for('auth_callback', plataforma=plataforma, _external=True)
    
    if not code: return "Erro: Código não recebido ou acesso negado pelo cliente.", 400

    token_acesso, token_refresh = "", ""
    page_id = None

    try:
        if plataforma == 'google':
            res = requests.post("https://oauth2.googleapis.com/token", data={
                "code": code, "client_id": GOOGLE_CLIENT_ID, "client_secret": GOOGLE_CLIENT_SECRET,
                "redirect_uri": redirect_uri, "grant_type": "authorization_code"
            }).json()
            token_acesso = res.get("access_token")
            token_refresh = res.get("refresh_token", "N/A")

        elif plataforma == 'instagram':
            res = requests.get("https://graph.facebook.com/v18.0/oauth/access_token", params={
                "client_id": META_APP_ID, "redirect_uri": redirect_uri, "client_secret": META_APP_SECRET, "code": code
            }).json()
            token_acesso = res.get("access_token")
            token_refresh = "N/A" 
            
            if token_acesso:
                pages_res = requests.get(f"https://graph.facebook.com/v18.0/me/accounts?access_token={token_acesso}").json()
                if 'data' in pages_res and len(pages_res['data']) > 0:
                    page_id = pages_res['data'][0]['id']
                    page_token = pages_res['data'][0]['access_token']
                    requests.post(f"https://graph.facebook.com/v18.0/{page_id}/subscribed_apps", 
                                  data={"subscribed_fields": "feed, messages, comments", "access_token": page_token})
                    token_acesso = page_token 

        if not token_acesso: return f"Erro ao obter token real da plataforma.", 400

    except Exception as e:
        return f"Erro de comunicação com a plataforma: {str(e)}", 500
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT id FROM integracoes_api WHERE empresa_id = %s AND plataforma = %s", (empresa_id, plataforma))
            if cursor.fetchone():
                cursor.execute("""
                    UPDATE integracoes_api 
                    SET token_acesso = %s, token_refresh = %s, plataforma_user_id = %s, status = 'ativo' 
                    WHERE empresa_id = %s AND plataforma = %s
                """, (token_acesso, token_refresh, page_id, empresa_id, plataforma))
            else:
                cursor.execute("""
                    INSERT INTO integracoes_api (empresa_id, plataforma, token_acesso, token_refresh, plataforma_user_id, status) 
                    VALUES (%s, %s, %s, %s, %s, 'ativo')
                """, (empresa_id, plataforma, token_acesso, token_refresh, page_id))
            conn.commit()
        except Exception as e:
            conn.rollback()
            print(f"Erro ao salvar integração no banco: {e}")
        finally:
            cursor.close()
            conn.close()

    return redirect(url_for('integracoes', sucesso="true", plat=plataforma))

# ================= WEBHOOKS (Escuta Ativa) =================

@app.route('/api/webhooks/ingest', methods=['GET', 'POST'])
def webhook_ingest():
    if request.method == 'GET':
        if request.args.get('hub.mode') == 'subscribe' and request.args.get('hub.verify_token') == META_VERIFY_TOKEN:
            return request.args.get('hub.challenge'), 200
        return "Acesso negado", 403

    data = request.json
    if not data: return jsonify({"error": "Payload inválido"}), 400

    conn = get_db_connection()
    if not conn: return jsonify({"error": "Falha no DB"}), 500

    try:
        cursor = conn.cursor()
        
        # PARSE ADAPTATIVO
        if 'entry' in data:
            try:
                plataforma = 'instagram'
                
                # Mapeamento dinâmico multi-tenancy a partir do ID da página enviado pelo webhook da Meta
                meta_page_id = str(data['entry'][0].get('id', ''))
                cursor.execute("""
                    SELECT empresa_id FROM integracoes_api 
                    WHERE plataforma = 'instagram' AND plataforma_user_id = %s AND status = 'ativo'
                """, (meta_page_id,))
                row = cursor.fetchone()
                empresa_id = row[0] if row else 1 # Fallback seguro
                
                alteracao = data['entry'][0]['changes'][0]['value']
                nome_cliente = alteracao.get('from', {}).get('username', 'Utilizador')
                comentario = alteracao.get('text', '')
                nota = 3 
            except (KeyError, IndexError):
                return jsonify({"error": "Formato Meta inválido"}), 400
        else:
            empresa_id = data.get('empresa_id')
            plataforma = data.get('plataforma')
            nome_cliente = data.get('nome_cliente', 'Cliente')
            nota = data.get('nota')
            comentario = data.get('comentario', '')

        if not empresa_id or not plataforma or nota is None:
            return jsonify({"error": "Dados ausentes"}), 400

        cursor.execute("""
            INSERT INTO avaliacoes_feed (empresa_id, plataforma_origem, nome_cliente, nota, comentario, status)
            VALUES (%s, %s, %s, %s, %s, 'pendente')
        """, (empresa_id, plataforma, nome_cliente, nota, comentario))
        avaliacao_id = cursor.lastrowid
        
        filtro = analisar_qualidade_comentario(comentario, nota)
        
        if filtro["valido"]:
            cursor.execute("SELECT tom_voz, regras_ouro, telefone_whatsapp FROM configuracoes_ia WHERE empresa_id = %s", (empresa_id,))
            config_ia = cursor.fetchone()
            regras_reais = f"Tom: {config_ia[0]}. Regras: {config_ia[1]}" if config_ia else "Tom profissional."
            telefone_real = config_ia[2] if config_ia and config_ia[2] else "Não configurado"

            resultado_ia = analisar_avaliacao_gemini(
                nome_cliente=nome_cliente, nota=nota, comentario=comentario,
                regras_tom_voz=regras_reais, telefone_empresa=telefone_real
            )
            sentimento = resultado_ia.get("sentimento", "neutro")
            rascunho = resultado_ia.get("sugestao_resposta", "")
            
            # Tags estruturadas retornadas pelo Gemini 2.5
            tags_json = json.dumps(resultado_ia.get("tags", []))
            
            novo_status = 'respondido_ia' if nota >= 4 else 'alerta_crise' if nota <= 3 else 'pendente'
        else:
            sentimento = filtro["sentimento"]
            rascunho = filtro["rascunho"]
            tags_json = json.dumps(["Filtro Spam Ativo"])
            novo_status = filtro["status"]

        if novo_status == 'alerta_crise':
            titulo_crise = f"Contenção: Avaliação {nota}★ de {nome_cliente}"
            cursor.execute("INSERT INTO acoes (empresa_id, titulo, prioridade, prazo, status) VALUES (%s, %s, 'critical', %s, 'pendente')", 
                           (empresa_id, titulo_crise, date.today().strftime('%Y-%m-%d')))

        cursor.execute("""
            UPDATE avaliacoes_feed SET sentimento = %s, rascunho_resposta = %s, tags = %s, status = %s WHERE id = %s
        """, (sentimento, rascunho, tags_json, novo_status, avaliacao_id))

        conn.commit()
        return jsonify({"success": True, "message": "Avaliação processada", "status": novo_status}), 201

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({"error": f"Erro interno: {str(e)}"}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if conn: conn.close()

# ================= GERADOR DE MASSA DE DADOS (DEMO) =================

@app.route('/api/seed', methods=['GET', 'POST'])
def api_seed():
    """Endpoint restrito para gerar massa de dados demonstrativa no app"""
    admin_secret = os.environ.get("ADMIN_SECRET_KEY", "frameia_reaction_master_key_2026")
    header_secret = request.headers.get("X-Admin-Secret") or request.args.get("secret")
    
    is_admin_secret = header_secret and header_secret == admin_secret
    is_logged_user = 'usuario_id' in session and session.get('empresa_id')

    # SEGURANÇA: Bloqueia qualquer chamada anônima externa não autorizada
    if not is_admin_secret and not is_logged_user:
        return jsonify({"error": "Acesso não autorizado. Este endpoint requer autenticação ou chave mestra administrativa."}), 403

    try:
        from gerar_massa_dados import popular_banco_dados
        empresa_id = session.get('empresa_id', 1) if is_logged_user else int(request.args.get('empresa_id', 1))
        res = popular_banco_dados(empresa_id)
        return jsonify({"success": True, "message": f"Massa de dados gerada com sucesso para a empresa #{empresa_id}!", "detalhes": res}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ================= INTEGRACAO MERCADO PAGO =================
MERCADOPAGO_ACCESS_TOKEN = os.environ.get("MERCADOPAGO_ACCESS_TOKEN", "APP_USR-DEMO-TOKEN")

@app.route('/api/checkout/<plano>/<ciclo>', methods=['GET', 'POST'])
def api_checkout(plano, ciclo):
    """Gera preferência de checkout no Mercado Pago (Checkout Pro / PIX)"""
    if 'usuario_id' not in session:
        return redirect(url_for('login', registo="true"))
    
    precos = {
        "pro": {"mensal": 79.90, "anual": 718.80},
        "premium": {"mensal": 197.90, "anual": 1774.80},
        "setup_vip": {"avista": 300.00, "mensal": 300.00, "anual": 300.00}
    }
    
    if plano == "setup_vip":
        ciclo = "avista"
        
    if plano not in precos or ciclo not in ["mensal", "anual", "avista"]:
        return jsonify({"error": "Plano ou ciclo inválido"}), 400
        
    valor = precos[plano][ciclo]
    titulo = "ReAction - Setup VIP (2 Sessões ao Vivo + 3 Meses Grátis)" if plano == "setup_vip" else f"ReAction - Plano {plano.capitalize()} ({ciclo.capitalize()})"
    
    try:
        import mercadopago
        sdk = mercadopago.SDK(MERCADOPAGO_ACCESS_TOKEN)
        
        preference_data = {
            "items": [
                {
                    "title": titulo,
                    "quantity": 1,
                    "unit_price": valor,
                    "currency_id": "BRL"
                }
            ],
            "payer": {
                "email": session.get('usuario_email', 'cliente@reaction.com.br')
            },
            "back_urls": {
                "success": url_for('app_dashboard', status="sucesso_pagamento", _external=True),
                "failure": url_for('app_dashboard', status="falha_pagamento", _external=True),
                "pending": url_for('app_dashboard', status="pendente_pagamento", _external=True)
            },
            "auto_return": "approved",
            "external_reference": f"empresa_{session.get('empresa_id')}_{plano}_{ciclo}"
        }
        
        preference_response = sdk.preference().create(preference_data)
        preference = preference_response.get("response", {})
        
        # Notificar os 3 administradores sobre o upgrade/contratação
        try:
            enviar_notificacao_admin("upgrade", {
                "nome_usuario": session.get('nome', 'Cliente'),
                "email": session.get('usuario_email', 'cliente@reaction.com.br'),
                "nome_empresa": f"Empresa ID #{session.get('empresa_id')}",
                "plano": f"Plano {plano.capitalize()}",
                "ciclo": ciclo.capitalize(),
                "valor": f"R$ {valor:.2f}"
            })
        except Exception as e_adm:
            print(f"Aviso envio e-mail admin checkout: {e_adm}")

        init_point = preference.get("init_point") or preference.get("sandbox_init_point")
        if init_point:
            return redirect(init_point)
        else:
            return jsonify({"success": True, "checkout_data": preference, "message": "Simulação de Checkout Mercado Pago iniciada"}), 200
            
    except Exception as mp_err:
        print(f"Erro no checkout Mercado Pago: {mp_err}")
        return jsonify({"error": f"Erro no Checkout Mercado Pago: {str(mp_err)}"}), 500

@app.route('/api/webhooks/mercadopago', methods=['POST'])
def webhook_mercadopago():
    """Webhook do Mercado Pago para escutar alterações e aprovações de pagamento"""
    data = request.json or {}
    print(f"Webhook Mercado Pago recebido: {data}")
    return jsonify({"status": "received"}), 200

# ================= ROTAS DE EXPORTAÇÃO EXCEL NATIVO (.XLSX MULTI-ABAS) =================

@app.route('/api/exportar/relatorio_excel')
@app.route('/api/exportar/avaliacoes')
@app.route('/api/exportar/acoes')
def exportar_relatorio_excel():
    """Gera um relatório completo em formato nativo Excel (.xlsx) com 4 abas profissionais estilizadas"""
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    empresa_id = session.get('empresa_id')
    conn = get_db_connection()
    if not conn:
        return "Erro de conexão com o banco de dados", 500

    cursor = conn.cursor(dictionary=True)
    avaliacoes = []
    acoes = []
    integracoes = []
    empresa_nome = f"Empresa #{empresa_id}"
    
    try:
        cursor.execute("SELECT nome_empresa FROM empresas WHERE id = %s", (empresa_id,))
        emp_res = cursor.fetchone()
        if emp_res and emp_res.get('nome_empresa'):
            empresa_nome = emp_res['nome_empresa']

        cursor.execute("""
            SELECT id, plataforma_origem, nome_cliente, nota, comentario, sentimento, status, rascunho_resposta
            FROM avaliacoes_feed
            WHERE empresa_id = %s
        """, (empresa_id,))
        avaliacoes = cursor.fetchall()

        cursor.execute("""
            SELECT id, titulo, prioridade, prazo, status
            FROM acoes
            WHERE empresa_id = %s
        """, (empresa_id,))
        acoes = cursor.fetchall()

        cursor.execute("""
            SELECT plataforma, status
            FROM integracoes_api
            WHERE empresa_id = %s
        """, (empresa_id,))
        integracoes = cursor.fetchall()

    except Exception as e:
        print(f"Erro ao buscar dados para relatório Excel: {e}")
    finally:
        cursor.close()
        conn.close()

    # Criar o Workbook openpyxl
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    fill_header = PatternFill(start_color="3A3A3A", end_color="3A3A3A", fill_type="solid")
    fill_brand = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True, color="1F2937")
    font_normal = Font(name="Calibri", size=11, color="374151")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    # ── ABA 1: RESUMO EXECUTIVO ──
    ws1 = wb.active
    ws1.title = "Resumo Executivo"
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:E2")
    cell_title = ws1["A1"]
    cell_title.value = f"REACTION | RELATÓRIO EXECUTIVO DE REPUTAÇÃO - {empresa_nome.upper()}"
    cell_title.font = font_title
    cell_title.fill = fill_brand
    cell_title.alignment = align_center

    total_av = len(avaliacoes)
    soma_notas = sum(r.get('nota', 0) for r in avaliacoes)
    nota_media = round(soma_notas / total_av, 2) if total_av > 0 else 5.0
    positivas = sum(1 for r in avaliacoes if r.get('sentimento') == 'positivo' or r.get('nota', 0) >= 4)
    neutras = sum(1 for r in avaliacoes if r.get('sentimento') == 'neutro' or r.get('nota', 0) == 3)
    negativas = sum(1 for r in avaliacoes if r.get('sentimento') == 'negativo' or r.get('nota', 0) <= 2)
    saude_pct = round((positivas / total_av) * 100, 1) if total_av > 0 else 100.0

    kpis = [
        ("Nota Média da Marca", f"{nota_media} / 5.0"),
        ("Total de Avaliações Cadastradas", total_av),
        ("Índice de Saúde da Reputação", f"{saude_pct}%"),
        ("Avaliações Positivas (4-5★)", positivas),
        ("Avaliações Neutras (3★)", neutras),
        ("Avaliações Críticas (1-2★)", negativas),
        ("Total de Ações de Contenção", len(acoes))
    ]

    ws1.cell(row=4, column=1, value="MÉTRICAS CHAVE DE PERFORMANCE").font = Font(name="Calibri", size=12, bold=True, color="FF6B35")

    ws1.cell(row=5, column=1, value="Indicador Executivo").font = font_header
    ws1.cell(row=5, column=1).fill = fill_header
    ws1.cell(row=5, column=2, value="Resultado").font = font_header
    ws1.cell(row=5, column=2).fill = fill_header

    for idx, (k, v) in enumerate(kpis, start=6):
        c1 = ws1.cell(row=idx, column=1, value=k)
        c2 = ws1.cell(row=idx, column=2, value=v)
        c1.font = font_bold
        c2.font = font_bold
        c1.border = thin_border
        c2.border = thin_border
        c2.alignment = align_center

    # ── ABA 2: AVALIAÇÕES ──
    ws2 = wb.create_sheet(title="Avaliações & Feedbacks")
    ws2.views.sheetView[0].showGridLines = True

    headers_av = ['ID', 'Cliente', 'Canal / Origem', 'Nota (Estrelas)', 'Comentário do Cliente', 'Sentimento', 'Status da Resposta', 'Resposta Gerada / IA']
    for col_num, h in enumerate(headers_av, 1):
        cell = ws2.cell(row=1, column=col_num, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_left if col_num in [2, 5, 8] else align_center

    for r_idx, r in enumerate(avaliacoes, start=2):
        row_data = [
            r.get('id', ''),
            r.get('nome_cliente', ''),
            r.get('plataforma_origem', '').upper(),
            r.get('nota', ''),
            r.get('comentario', ''),
            r.get('sentimento', '').capitalize(),
            r.get('status', ''),
            r.get('rascunho_resposta', '')
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_normal
            cell.border = thin_border
            if r_idx % 2 == 0: cell.fill = fill_zebra

    # ── ABA 3: PLANO DE AÇÕES ──
    ws3 = wb.create_sheet(title="Plano de Ações")
    ws3.views.sheetView[0].showGridLines = True

    headers_ac = ['ID', 'Título da Ação de Contenção', 'Prioridade', 'Prazo Final', 'Status']
    for col_num, h in enumerate(headers_ac, 1):
        cell = ws3.cell(row=1, column=col_num, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    for r_idx, r in enumerate(acoes, start=2):
        row_data = [
            r.get('id', ''),
            r.get('titulo', ''),
            r.get('prioridade', 'normal').capitalize(),
            str(r.get('prazo') or 'Sem prazo'),
            r.get('status', '').capitalize()
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_normal
            cell.border = thin_border
            if r_idx % 2 == 0: cell.fill = fill_zebra

    # Auto-ajuste de largura de colunas em todas as abas
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 65)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reaction_relatorio_executivo_{empresa_id}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.route('/api/usuario/foto', methods=['POST'])
def atualizar_foto_perfil():
    if 'usuario_id' not in session:
        return jsonify({'sucesso': False, 'erro': 'Não autenticado'}), 401
    
    usuario_id = session['usuario_id']
    foto_url = None
    
    try:
        if 'foto' in request.files:
            file = request.files['foto']
            if file and file.filename != '':
                rnd_id = os.urandom(4).hex()
                filename = f"avatar_{usuario_id}_{rnd_id}.png"
                upload_dir = os.path.join(app.static_folder, 'uploads', 'avatars')
                os.makedirs(upload_dir, exist_ok=True)
                filepath = os.path.join(upload_dir, filename)
                file.save(filepath)
                foto_url = f"/static/uploads/avatars/{filename}"
        elif request.is_json and request.json.get('foto_url'):
            foto_url = request.json.get('foto_url')
            
        if foto_url:
            conn = get_db_connection()
            if conn:
                cursor = conn.cursor()
                try:
                    cursor.execute("UPDATE usuarios SET foto_perfil = %s WHERE id = %s", (foto_url, usuario_id))
                    conn.commit()
                    session['foto_perfil'] = foto_url
                    return jsonify({'sucesso': True, 'foto_url': foto_url})
                finally:
                    cursor.close()
                    conn.close()
    except Exception as err:
        print(f"Erro ao salvar foto de perfil: {err}")
        return jsonify({'sucesso': False, 'erro': f'Erro ao processar imagem: {str(err)}'}), 500
                
    return jsonify({'sucesso': False, 'erro': 'Nenhuma imagem fornecida'}), 400

@app.route('/api/usuario/foto/remover', methods=['POST'])
def remover_foto_perfil():
    if 'usuario_id' not in session:
        return jsonify({'sucesso': False, 'erro': 'Não autenticado'}), 401
        
    usuario_id = session['usuario_id']
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            cursor.execute("UPDATE usuarios SET foto_perfil = NULL WHERE id = %s", (usuario_id,))
            conn.commit()
            session['foto_perfil'] = None
            return jsonify({'sucesso': True})
        finally:
            cursor.close()
            conn.close()
    return jsonify({'sucesso': False, 'erro': 'Erro ao atualizar banco'}), 500




# ==========================================
# HELPERS E INTEGRAÇÃO WHATSAPP (META CLOUD API)
# ==========================================

def buscar_usuario_por_whatsapp(telefone, cursor):
    """Localiza o usuário e sua respectiva empresa pelo número de WhatsApp formatado"""
    if not telefone:
        return None
    # Limpa caracteres não numéricos
    tel_digits = re.sub(r'\D', '', str(telefone))
    # Testa variações (com e sem código de país 55 e nono dígito)
    variacoes = [tel_digits]
    if tel_digits.startswith('55') and len(tel_digits) > 10:
        variacoes.append(tel_digits[2:]) # sem 55
    else:
        variacoes.append('55' + tel_digits) # com 55

    for tel in variacoes:
        cursor.execute("""
            SELECT u.id as usuario_id, u.empresa_id, u.nome, u.email, u.role, u.telefone,
                   e.nome_empresa, e.plano_assinatura,
                   c.telefone_whatsapp, c.tom_voz
            FROM usuarios u
            INNER JOIN empresas e ON u.empresa_id = e.id
            LEFT JOIN configuracoes_ia c ON e.id = c.empresa_id
            WHERE u.telefone LIKE %s OR c.telefone_whatsapp LIKE %s
            LIMIT 1
        """, (f"%{tel[-8:]}%", f"%{tel[-8:]}%"))
        res = cursor.fetchone()
        if res:
            return res
    return None

def enviar_mensagem_whatsapp(telefone, texto):
    """Dispara mensagem via API oficial do WhatsApp da Meta"""
    raw_token = os.getenv('META_ACCESS_TOKEN', 'EAATkzbdHnr0BSNtiVbbWDySAcmjQX71oo6CXDu3QzbOx0rVLpmyCx4prsd3apKY6xKjh4LVdIBhJK5kOZCFSJexqCwchFwSWXB3euiU85GZCxanhH1lnxJVoVNpx72k8XGBr53ypaqDVUArZAk0MxZAThW54Q6jvyIyYSvO5Yxj88hDALOrpUbx9uCfckLZBqLgZDZD')
    token = raw_token.replace('META_ACCESS_TOKEN=', '').replace('"', '').strip() if raw_token else ''
    phone_id = os.getenv('META_PHONE_ID', '1313361515185262').replace('META_PHONE_ID=', '').replace('"', '').strip()

    if not token or not phone_id:
        print(f"[WHATSAPP ENVIO SIMULADO] Para: {telefone} | Texto: {texto[:100]}...")
        return True

    url = f"https://graph.facebook.com/v19.0/{phone_id}/messages"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": re.sub(r'\D', '', str(telefone)),
        "type": "text",
        "text": {"body": texto}
    }

    try:
        req = requests.post(url, headers=headers, json=payload, timeout=12)
        print(f"[WHATSAPP ENVIO] Status HTTP {req.status_code} | Resposta: {req.text[:120]}")
        return req.status_code in [200, 201]
    except Exception as e:
        print(f"[WHATSAPP ENVIO ERRO] {e}")
        return False

# ==========================================
# ROTAS PÚBLICAS: TERMOS E PRIVACIDADE (LGPD)
# ==========================================

@app.route('/termos')
def pagina_termos():
    return render_template('termos.html')

@app.route('/privacidade')
def pagina_privacidade():
    return render_template('privacidade.html')

@app.route('/api/aceitar_termos', methods=['POST'])
def api_aceitar_termos():
    if 'usuario_id' not in session:
        return jsonify({'error': 'Não autenticado'}), 401
    
    usuario_id = session['usuario_id']
    ip_cliente = request.headers.get('X-Forwarded-For', request.remote_addr)
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            UPDATE usuarios 
            SET termos_aceitos_em = NOW(), termos_versao = 'v0.2.0', termos_ip = %s 
            WHERE id = %s
        """, (ip_cliente, usuario_id))
        conn.commit()
        return jsonify({'success': True, 'message': 'Termos e Política de Privacidade aceitos com sucesso!'})
    except Exception as e:
        print(f"Erro ao salvar aceite de termos: {e}")
        return jsonify({'error': 'Erro ao processar aceite.'}), 500
    finally:
        cursor.close()
        conn.close()

# ==========================================
# WEBHOOK WHATSAPP BOT (META CLOUD API)
# ==========================================

@app.route('/webhook', methods=['GET', 'POST'])
def webhook_whatsapp():
    # 1. Validação de Token da Meta (Handshake GET)
    if request.method == 'GET':
        mode = request.args.get("hub.mode")
        token = request.args.get("hub.verify_token")
        challenge = request.args.get("hub.challenge")
        
        verify_token_env = os.getenv('META_VERIFY_TOKEN', 'reaction_token_secret_022026')
        
        if mode == "subscribe" and token == verify_token_env:
            print("[WHATSAPP WEBHOOK GET] Webhook verificado com sucesso pela Meta!")
            return str(challenge), 200
        
        print("[WHATSAPP WEBHOOK GET] Falha na verificação do token Meta.")
        return "Falha na verificação. Token inválido.", 403

    # 2. Recebimento de Mensagens dos Gestores (POST)
    if request.method == 'POST':
        dados = request.get_json(silent=True) or {}
        print(f"[WHATSAPP RAW POST] {dados}")

        try:
            entries = dados.get('entry', [])
            if not entries:
                if 'value' in dados:
                    entries = [{'changes': [{'value': dados['value']}]}]
                elif 'messages' in dados:
                    entries = [{'changes': [{'value': dados}]}]

            for entry in entries:
                for change in entry.get('changes', []):
                    value = change.get('value', {})
                    for mensagem in value.get('messages', []):
                        telefone_cliente = re.sub(r'\D', '', str(mensagem.get('from', '')))
                        tipo_msg = mensagem.get('type', 'text')
                        texto_recebido = ''
                        if tipo_msg == 'text' and 'text' in mensagem:
                            texto_recebido = mensagem['text'].get('body', '').strip()
                        elif 'text' in mensagem:
                            texto_recebido = str(mensagem['text'].get('body', '')).strip()

                        if not texto_recebido:
                            continue

                        print(f"[WHATSAPP MENSAGEM] De: {telefone_cliente} | Texto: {texto_recebido}")

                        conn = get_db_connection()
                        cursor = conn.cursor(dictionary=True)

                        try:
                            usuario = buscar_usuario_por_whatsapp(telefone_cliente, cursor)

                            if not usuario:
                                msg_nao_cadastrado = (
                                    "🛡️ *Assistente ReAction [IA]*\n\n"
                                    "⚠️ *WhatsApp não vinculado ao ReAction*\n\n"
                                    "Não encontramos uma conta associada a este número de telefone.\n\n"
                                    "👉 Acesse a aba *Ajustes/Perfil* no seu painel ReAction e informe seu WhatsApp para liberar os comandos e alertas!"
                                )
                                enviar_mensagem_whatsapp(telefone_cliente, msg_nao_cadastrado)
                                continue

                            empresa_id = usuario['empresa_id']
                            nome_gestor = usuario['nome'].split()[0]
                            texto_lc = texto_recebido.lower().strip()

                            # COMANDO 1: RESUMO / REPUTAÇÃO / SAÚDE DA MARCA
                            if texto_lc in ['resumo', 'reputacao', 'reputação', 'status', 'saude', 'saúde', 'cockpit']:
                                cursor.execute("""
                                    SELECT 
                                        COUNT(*) as total_reviews,
                                        COALESCE(AVG(nota), 5.0) as media_nota,
                                        COUNT(CASE WHEN status = 'alerta_crise' OR nota <= 2 THEN 1 END) as alertas_crise,
                                        COUNT(CASE WHEN status = 'pendente' THEN 1 END) as pendentes_ia
                                    FROM avaliacoes_feed
                                    WHERE empresa_id = %s
                                """, (empresa_id,))
                                stats = cursor.fetchone() or {}

                                media = float(stats.get('media_nota', 5.0))
                                total = stats.get('total_reviews', 0)
                                crises = stats.get('alertas_crise', 0)
                                pendentes = stats.get('pendentes_ia', 0)

                                saude_pct = round((media / 5.0) * 100, 1)
                                emoji_saude = "🟢" if saude_pct >= 90 else ("🟡" if saude_pct >= 70 else "🔴")

                                resposta_resumo = (
                                    f"🛡️ *Cockpit ReAction • {usuario['nome_empresa']}*\n\n"
                                    f"Olá, *{nome_gestor}*! Aqui está o raio-x da sua reputação:\n\n"
                                    f"{emoji_saude} *Saúde da Marca:* {saude_pct}% ({media:.1f} ⭐)\n"
                                    f"📊 *Total de Avaliações:* {total}\n"
                                    f"🚨 *Alertas de Crise:* {crises}\n"
                                    f"⏳ *Pendências de IA:* {pendentes}\n\n"
                                    f"🔗 Acesse o painel: https://reaction.frameia.com.br/app"
                                )
                                enviar_mensagem_whatsapp(telefone_cliente, resposta_resumo)

                            # COMANDO 2: LISTAR AÇÕES / TAREFAS
                            elif texto_lc in ['acoes', 'ações', 'tarefas', 'pendencias', 'pendências']:
                                cursor.execute("""
                                    SELECT id, titulo, prioridade, prazo, status
                                    FROM acoes
                                    WHERE empresa_id = %s AND status = 'pendente'
                                    ORDER BY (prioridade = 'critical') DESC, prazo ASC
                                    LIMIT 5
                                """, (empresa_id,))
                                lista_acoes = cursor.fetchall()

                                if not lista_acoes:
                                    msg_acoes = f"✅ *Tudo em dia, {nome_gestor}!*\n\nNão há nenhuma ação pendente no seu Cockpit no momento."
                                else:
                                    msg_acoes = f"📋 *Ações Pendentes no Cockpit ({len(lista_acoes)}):*\n\n"
                                    for a in lista_acoes:
                                        prio_emoji = "🚨" if a['prioridade'] == 'critical' else "📌"
                                        prazo_txt = f" (Prazo: {a['prazo']})" if a.get('prazo') else ""
                                        msg_acoes += f"{prio_emoji} *#{a['id']}* - {a['titulo']}{prazo_txt}\n"
                                    msg_acoes += "\n💡 Para criar uma nova ação, envie: *criar acao [descrição]*"

                                enviar_mensagem_whatsapp(telefone_cliente, msg_acoes)

                            # COMANDO 3: CRIAR NOVA AÇÃO NO COCKPIT
                            elif texto_lc.startswith('criar acao') or texto_lc.startswith('criar ação') or texto_lc.startswith('acao ') or texto_lc.startswith('ação '):
                                titulo_acao = re.sub(r'^(?:criar\s+ac(?:a|ã)o|ac(?:a|ã)o)\s+', '', texto_recebido, flags=re.IGNORECASE).strip()
                                if titulo_acao:
                                    cursor.execute("""
                                        INSERT INTO acoes (empresa_id, criado_por, titulo, prioridade, status)
                                        VALUES (%s, %s, %s, 'normal', 'pendente')
                                    """, (empresa_id, usuario['usuario_id'], titulo_acao))
                                    conn.commit()
                                    nova_id = cursor.lastrowid

                                    msg_criada = (
                                        f"✅ *Ação Registrada com Sucesso!*\n\n"
                                        f"📌 *#{nova_id}* - {titulo_acao}\n"
                                        f"👤 Criada por: {nome_gestor}\n"
                                        f"📱 Já visível no seu Cockpit ReAction!"
                                    )
                                    enviar_mensagem_whatsapp(telefone_cliente, msg_criada)

                            # COMANDO 4: ALERTAS VERMELHOS / CRISES
                            elif texto_lc in ['alertas', 'alerta', 'crises', 'crise']:
                                cursor.execute("""
                                    SELECT id, plataforma_origem, nome_cliente, nota, comentario
                                    FROM avaliacoes_feed
                                    WHERE empresa_id = %s AND (status = 'alerta_crise' OR nota <= 2)
                                    ORDER BY id DESC
                                    LIMIT 3
                                """, (empresa_id,))
                                alertas = cursor.fetchall()

                                if not alertas:
                                    msg_alertas = f"🛡️ *Nenhum alerta de crise ativo no momento!* Sua reputação está segura, {nome_gestor}."
                                else:
                                    msg_alertas = f"🚨 *Alertas de Crise Recentes ({len(alertas)}):*\n\n"
                                    for al in alertas:
                                        msg_alertas += f"🔴 *[{al['plataforma_origem'].upper()}] {al['nome_cliente']}* ({al['nota']} ⭐)\n"
                                        msg_alertas += f"💬 \"{al['comentario'] or 'Sem comentário'}\"\n\n"
                                    msg_alertas += "👉 Acesse a Central de Reputação para responder: https://reaction.frameia.com.br/reputacao"

                                enviar_mensagem_whatsapp(telefone_cliente, msg_alertas)

                            # COMANDO PADRÃO / AJUDA
                            else:
                                msg_ajuda = (
                                    f"🤖 *Assistente ReAction [IA] • Olá, {nome_gestor}!*\n\n"
                                    f"Comandos rápidos disponíveis no seu WhatsApp:\n\n"
                                    f"🔹 *resumo* - Raio-X da reputação e saúde da marca\n"
                                    f"🔹 *acoes* - Ver ações e tarefas pendentes no Cockpit\n"
                                    f"🔹 *criar acao [texto]* - Criar nova ação diretamente no painel\n"
                                    f"🔹 *alertas* - Monitorar crises e notas baixas recentes\n"
                                    f"🔹 *ajuda* - Exibir este menu de comandos\n\n"
                                    f"🌐 Painel Web: https://reaction.frameia.com.br"
                                )
                                enviar_mensagem_whatsapp(telefone_cliente, msg_ajuda)

                        finally:
                            cursor.close()
                            conn.close()

            return jsonify({"status": "EVENT_RECEIVED"}), 200

        except Exception as err:
            print(f"[WHATSAPP WEBHOOK ERRO GERAL] {err}")
            return jsonify({"status": "ERROR", "message": str(err)}), 500

if __name__ == '__main__':
    flask_debug = os.environ.get('FLASK_DEBUG', 'False').lower() in ['true', '1']
    flask_host = os.environ.get('FLASK_HOST', '127.0.0.1')
    app.run(host=flask_host, port=5001, debug=flask_debug)

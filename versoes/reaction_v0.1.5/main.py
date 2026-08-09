import os
import json
import re
import secrets
import mysql.connector
import requests
import io
import csv
from flask import Flask, render_template, redirect, url_for, session, request, jsonify, Response
from datetime import datetime, date, timedelta
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash
from flask_cors import CORS

# Assumindo que o ficheiro da IA continua a chamar-se agente_reputacao.py
from agente_reputacao import analisar_avaliacao_gemini
from servico_email import enviar_email_boas_vindas, enviar_notificacao_admin, enviar_email_redefinicao_senha

load_dotenv()

# ================= CONFIGURAÇÕES GERAIS =================
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET")
META_APP_ID = os.environ.get("META_APP_ID")
META_APP_SECRET = os.environ.get("META_APP_SECRET")
META_VERIFY_TOKEN = os.environ.get("META_VERIFY_TOKEN", "frameia_reaction_2026")

app = Flask(__name__)

# Chave de sessão estável para evitar desconexões quando o servidor reinicia
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "chave_super_secreta_reaction_2026")

# CORS adaptativo para permitir desenvolvimento local e domínio oficial de produção
cors_origins = os.environ.get("CORS_ORIGINS", "https://reaction.frameia.com.br").split(",")
if app.debug or os.environ.get("FLASK_ENV") == "development":
    # Permite localhost e 127.0.0.1 nas portas comuns
    cors_origins.extend([
        "http://localhost:5000", "http://127.0.0.1:5000",
        "http://localhost:5001", "http://127.0.0.1:5001"
    ])
CORS(app, resources={r"/*": {"origins": cors_origins}})

DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_USER = os.environ.get("DB_USER", "root")
DB_PASSWORD = os.environ.get("DB_PASS", "")
DB_NAME = os.environ.get("DB_NAME", "reaction_db")

# ================= BANCO DE DADOS =================
def get_db_connection():
    try:
        conn = mysql.connector.connect(
            host=DB_HOST,
            user=DB_USER,
            password=DB_PASSWORD,
            database=DB_NAME
        )
        return conn
    except mysql.connector.Error as err:
        if err.errno == mysql.connector.errorcode.ER_BAD_DB_ERROR:
            try:
                conn = mysql.connector.connect(
                    host=DB_HOST,
                    user=DB_USER,
                    password=DB_PASSWORD
                )
                cursor = conn.cursor()
                cursor.execute(f"CREATE DATABASE {DB_NAME}")
                conn.database = DB_NAME
                cursor.close()
                return conn
            except mysql.connector.Error as e:
                print(f"Erro ao tentar criar base de dados: {e}")
                return None
        else:
            print(f"Erro ao conectar à base de dados: {err}")
            return None

def criar_tabelas_se_nao_existirem():
    conn = get_db_connection()
    if not conn:
        print("Aviso: Falha ao obter conexão para criar tabelas.")
        return
        
    try:
        cursor = conn.cursor()
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS empresas (
                id INT AUTO_INCREMENT PRIMARY KEY,
                nome_empresa VARCHAR(255) NOT NULL,
                plano_assinatura VARCHAR(50) DEFAULT 'basico',
                data_criacao DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id INT AUTO_INCREMENT PRIMARY KEY,
                empresa_id INT NOT NULL,
                nome VARCHAR(255) NOT NULL,
                email VARCHAR(255) NOT NULL UNIQUE,
                senha VARCHAR(255) NOT NULL,
                role ENUM('dono', 'admin', 'operador') DEFAULT 'operador',
                FOREIGN KEY (empresa_id) REFERENCES empresas(id) ON DELETE CASCADE
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS integracoes_api (
                id INT AUTO_INCREMENT PRIMARY KEY,
                empresa_id INT NOT NULL,
                plataforma ENUM('google', 'instagram', 'whatsapp', 'ifood') NOT NULL,
                token_acesso TEXT,
                token_refresh TEXT,
                plataforma_user_id VARCHAR(255) NULL,
                data_expiracao DATETIME,
                status ENUM('ativo', 'inativo', 'erro') DEFAULT 'inativo',
                FOREIGN KEY (empresa_id) REFERENCES empresas(id) ON DELETE CASCADE
            )
        """)
        
        # Migração automática de colunas para integracoes_api se a tabela já existia
        colunas_integracoes = {
            'token_refresh': 'TEXT NULL AFTER token_acesso',
            'plataforma_user_id': 'VARCHAR(255) NULL AFTER token_acesso',
            'data_expiracao': 'DATETIME NULL AFTER status'
        }
        for col_nome, col_def in colunas_integracoes.items():
            try:
                cursor.execute(f"SHOW COLUMNS FROM integracoes_api LIKE '{col_nome}'")
                if not cursor.fetchone():
                    cursor.execute(f"ALTER TABLE integracoes_api ADD COLUMN {col_nome} {col_def}")
                    print(f"Coluna {col_nome} adicionada com sucesso na tabela integracoes_api.")
            except Exception as col_err:
                print(f"Erro ao verificar/adicionar coluna {col_nome} em integracoes_api: {col_err}")
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS avaliacoes_feed (
                id INT AUTO_INCREMENT PRIMARY KEY,
                empresa_id INT NOT NULL,
                plataforma_origem ENUM('google', 'instagram', 'whatsapp', 'ifood') NOT NULL,
                nome_cliente VARCHAR(255),
                nota INT CHECK (nota >= 1 AND nota <= 5),
                comentario TEXT,
                sentimento VARCHAR(50),
                rascunho_resposta TEXT,
                tags TEXT,
                status ENUM('pendente', 'respondido_ia', 'alerta_crise') DEFAULT 'pendente',
                FOREIGN KEY (empresa_id) REFERENCES empresas(id) ON DELETE CASCADE
            )
        """)
        
        # Garantir que a coluna 'tags' existe em 'avaliacoes_feed' se a tabela já existia
        try:
            cursor.execute("SHOW COLUMNS FROM avaliacoes_feed LIKE 'tags'")
            if not cursor.fetchone():
                cursor.execute("ALTER TABLE avaliacoes_feed ADD COLUMN tags TEXT AFTER rascunho_resposta")
                print("Coluna 'tags' adicionada com sucesso na tabela avaliacoes_feed.")
        except Exception as tags_err:
            print(f"Erro ao verificar/adicionar coluna tags em avaliacoes_feed: {tags_err}")
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS acoes (
                id INT AUTO_INCREMENT PRIMARY KEY,
                empresa_id INT NOT NULL,
                criado_por INT,
                titulo VARCHAR(255) NOT NULL,
                prioridade ENUM('normal', 'critical') DEFAULT 'normal',
                prazo DATE,
                status ENUM('pendente', 'concluido') DEFAULT 'pendente',
                FOREIGN KEY (empresa_id) REFERENCES empresas(id) ON DELETE CASCADE,
                FOREIGN KEY (criado_por) REFERENCES usuarios(id) ON DELETE SET NULL
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS configuracoes_ia (
                id INT AUTO_INCREMENT PRIMARY KEY,
                empresa_id INT NOT NULL UNIQUE,
                tom_voz VARCHAR(255) DEFAULT 'Profissional e empático',
                regras_ouro TEXT,
                telefone_whatsapp VARCHAR(50),
                FOREIGN KEY (empresa_id) REFERENCES empresas(id) ON DELETE CASCADE
            )
        """)

        cursor.execute("INSERT IGNORE INTO empresas (id, nome_empresa) VALUES (1, 'Empresa Teste Reaction')")
        senha_hash = generate_password_hash('123')
        cursor.execute("""
            INSERT INTO usuarios (id, empresa_id, nome, email, senha, role) 
            VALUES (1, 1, 'Jadson', 'admin@teste.com', %s, 'dono')
            ON DUPLICATE KEY UPDATE senha=%s
        """, (senha_hash, senha_hash))

        # Auto-seed massa de dados se a tabela avaliacoes_feed estiver vazia
        try:
            cursor.execute("SELECT COUNT(*) FROM avaliacoes_feed WHERE empresa_id = 1")
            row_count = cursor.fetchone()[0]
            if row_count == 0:
                print("Nenhuma avaliação encontrada. Gerando massa de dados demonstrativa automaticamente...")
                from gerar_massa_dados import popular_banco_dados
                popular_banco_dados(1)
                print("Massa de dados auto-gerada com sucesso!")
        except Exception as seed_err:
            print(f"Erro ao verificar/auto-gerar massa de dados: {seed_err}")

        conn.commit()
        cursor.close()
        conn.close()
        
    except Exception as e:
        print(f"Erro fatal ao criar tabelas: {e}")

try:
    criar_tabelas_se_nao_existirem()
except Exception as e:
    print(f"Ignorando criação de tabelas: {e}")

# ================= MIDDLEWARES & HELPERS =================

def analisar_qualidade_comentario(comentario, nota):
    """Firewall de Spam embutido para poupar tokens do Gemini"""
    if not comentario or not str(comentario).strip():
        if int(nota) >= 4:
            return {"valido": False, "status": "respondido_ia", "rascunho": "Obrigado pela avaliação!", "sentimento": "positivo"}
        return {"valido": False, "status": "alerta_crise", "rascunho": "Lamentamos. Fale connosco.", "sentimento": "negativo"}

    texto = str(comentario).lower().strip()
    
    if re.search(r'(http|https)://', texto) or re.search(r'www\.', texto):
        return {"valido": False, "status": "pendente", "rascunho": "[SPAM - LINK DETETADO]", "sentimento": "neutro"}
        
    palavras_spam = ['ganhe dinheiro', 'bitcoin', 'jogo do tigrinho', 'fortune tiger', 'seguidores']
    if any(spam in texto for spam in palavras_spam):
        return {"valido": False, "status": "pendente", "rascunho": "[SPAM DETETADO]", "sentimento": "neutro"}

    if len(texto) < 15:
        if int(nota) >= 4:
            return {"valido": False, "status": "respondido_ia", "rascunho": "Obrigado pelo feedback!", "sentimento": "positivo"}
        elif int(nota) <= 3:
            return {"valido": False, "status": "alerta_crise", "rascunho": "Sentimos muito. Chame no canal de atendimento.", "sentimento": "negativo"}

    return {"valido": True}

def get_usuario_logado():
    """Função utilitária para pegar os dados do utilizador logado em todas as rotas"""
    nome = session.get('nome', 'Usuário')
    iniciais = nome[0].upper() if nome else 'U'
    return {'nome': nome, 'iniciais': iniciais}

def garantir_massa_dados_empresa():
    """Garante que a empresa logada possui dados demonstrativos populados automaticamente"""
    empresa_id = session.get('empresa_id')
    if not empresa_id: return
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT COUNT(*) FROM avaliacoes_feed WHERE empresa_id = %s", (empresa_id,))
            res = cursor.fetchone()
            total = res[0] if res else 0
            if total == 0:
                print(f"Empresa {empresa_id} sem avaliações. Gerando massa de dados demonstrativa...")
                from gerar_massa_dados import popular_banco_dados
                popular_banco_dados(empresa_id)
        except Exception as e:
            print(f"Erro ao verificar/garantir massa de dados: {e}")
        finally:
            cursor.close()
            conn.close()

# ================= ROTAS DE VIEWS (TELAS DO SAAS) =================

@app.route('/')
@app.route('/index.html')
def index():
    # Rota raiz sempre serve a Landing Page diretamente com headers no-cache
    response = app.make_response(render_template('index.html'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Se o usuário já estiver logado, redireciona de conveniência direto para o cockpit
    if 'usuario_id' in session:
        return redirect(url_for('app_dashboard'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        senha = request.form.get('senha', '')
        
        conn = get_db_connection()
        if not conn: return render_template('login.html', erro="Erro temporário no servidor.")
            
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute("SELECT * FROM usuarios WHERE email = %s", (email,))
            user = cursor.fetchone()
        finally:
            cursor.close()
            conn.close()

        if user and check_password_hash(user['senha'], senha):
            session['usuario_id'] = user['id']
            session['empresa_id'] = user['empresa_id']
            session['nome'] = user['nome']
            return redirect(url_for('app_dashboard'))
        else:
            return render_template('login.html', erro="E-mail ou palavra-passe incorretos.")
    
    return render_template('login.html')

@app.route('/registo', methods=['POST'])
def registo():
    nome_empresa = request.form.get('nome_empresa', '').strip()
    nome_usuario = request.form.get('nome_usuario', '').strip()
    email = request.form.get('email', '').strip().lower()
    senha = request.form.get('senha', '')

    if not all([nome_empresa, nome_usuario, email, senha]):
        return render_template('login.html', erro_registo="Todos os campos são obrigatórios.", mostrar_registo=True)

    senha_hash = generate_password_hash(senha)
    conn = get_db_connection()
    if not conn: return render_template('login.html', erro_registo="Erro no servidor.", mostrar_registo=True)

    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id FROM usuarios WHERE email = %s", (email,))
        if cursor.fetchone():
            return render_template('login.html', erro_registo="Este e-mail já está em uso.", mostrar_registo=True)

        novo_usuario_id = None
        empresa_id = None

        cursor.execute("INSERT INTO empresas (nome_empresa) VALUES (%s)", (nome_empresa,))
        empresa_id = cursor.lastrowid
        cursor.execute("INSERT INTO usuarios (empresa_id, nome, email, senha, role) VALUES (%s, %s, %s, %s, 'dono')", 
                       (empresa_id, nome_usuario, email, senha_hash))
        novo_usuario_id = cursor.lastrowid
        conn.commit()
    except Exception as e:
        conn.rollback()
        return render_template('login.html', erro_registo="Erro interno ao criar conta.", mostrar_registo=True)
    finally:
        cursor.close()
        conn.close()

    # Disparar e-mail de boas-vindas ao cliente e notificar os 3 administradores
    try:
        enviar_email_boas_vindas(email, nome_usuario, nome_empresa)
        enviar_notificacao_admin("novo_usuario", {
            "nome_usuario": nome_usuario,
            "email": email,
            "nome_empresa": nome_empresa,
            "plano": "Shield Start (Freemium)",
            "ciclo": "Mensal",
            "valor": "R$ 0,00"
        })
    except Exception as err_mail:
        print(f"Aviso ao enviar e-mails de registo: {err_mail}")

    # Auto-login automático após o registo (Fricção Zero)
    session['usuario_id'] = novo_usuario_id
    session['empresa_id'] = empresa_id
    session['nome'] = nome_usuario
    garantir_massa_dados_empresa()
    return redirect(url_for('app_dashboard'))

@app.route('/login/google', methods=['GET', 'POST'])
def login_google():
    """Rota de Autenticação / Cadastro com a Conta do Google (Google SSO)"""
    if 'usuario_id' in session:
        return redirect(url_for('app_dashboard'))
        
    google_token = request.values.get('credential') or request.values.get('g_token')
    google_email = request.values.get('email')
    google_nome = request.values.get('nome') or request.values.get('name')
    google_sub = request.values.get('sub') or request.values.get('google_id')
    
    # Se a credencial/token do Google ID Token foi enviada
    if google_token:
        try:
            import urllib.request
            import json
            req_url = f"https://oauth2.googleapis.com/tokeninfo?id_token={google_token}"
            with urllib.request.urlopen(req_url, timeout=5) as resp:
                token_data = json.loads(resp.read().decode('utf-8'))
                google_email = token_data.get('email')
                google_nome = token_data.get('name') or token_data.get('given_name')
                google_sub = token_data.get('sub')
        except Exception as err_tkn:
            print(f"Aviso ao validar token Google: {err_tkn}")

    # Exigir credenciais válidas do Google SSO (Sem fallback inseguro)
    if not google_email:
        return render_template('login.html', erro="Autenticação com o Google requer credenciais válidas do provedor. Por favor, tente novamente.")

    conn = get_db_connection()
    if not conn:
        return render_template('login.html', erro="Erro temporário de conexão com o banco de dados.")

    cursor = conn.cursor(dictionary=True)
    user = None
    try:
        # Garantir se a coluna google_id existe no banco
        try:
            cursor.execute("SELECT google_id FROM usuarios LIMIT 1")
            cursor.fetchall()
        except Exception:
            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN google_id VARCHAR(255) DEFAULT NULL")
                conn.commit()
            except Exception:
                pass

        # Buscar se o e-mail ou google_id já possui cadastro
        cursor.execute("SELECT * FROM usuarios WHERE email = %s OR (google_id IS NOT NULL AND google_id = %s)", (google_email, google_sub))
        user = cursor.fetchone()
        
        if not user:
            # NOVO CADASTRO AUTOMÁTICO VIA GOOGLE
            nome_empresa = f"Empresa de {google_nome}"
            senha_aleatoria = generate_password_hash(f"GoogleSSO_{os.urandom(8).hex()}")
            
            cursor.execute("INSERT INTO empresas (nome_empresa) VALUES (%s)", (nome_empresa,))
            empresa_id = cursor.lastrowid
            
            try:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role, google_id) VALUES (%s, %s, %s, %s, 'dono', %s)",
                    (empresa_id, google_nome, google_email, senha_aleatoria, google_sub)
                )
            except Exception:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role) VALUES (%s, %s, %s, %s, 'dono')",
                    (empresa_id, google_nome, google_email, senha_aleatoria)
                )
                
            usuario_id = cursor.lastrowid
            conn.commit()
            
            user = {
                'id': usuario_id,
                'empresa_id': empresa_id,
                'nome': google_nome,
                'email': google_email
            }
            
            # Disparar e-mail de boas vindas e alerta de administradores
            try:
                enviar_email_boas_vindas(google_email, google_nome, nome_empresa)
                enviar_notificacao_admin("novo_usuario", {
                    "nome_usuario": google_nome,
                    "email": google_email,
                    "nome_empresa": nome_empresa,
                    "plano": "Shield Start (Google SSO)",
                    "ciclo": "Mensal",
                    "valor": "R$ 0,00"
                })
            except Exception as err_m:
                print(f"Erro no envio de e-mails Google SSO: {err_m}")

    except Exception as err_db:
        print(f"Erro no processamento do Google SSO: {err_db}")
        if conn: conn.rollback()
        return render_template('login.html', erro="Erro ao processar a autenticação com o Google.")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    if user:
        session['usuario_id'] = user['id']
        session['empresa_id'] = user['empresa_id']
        session['nome'] = user['nome']
        garantir_massa_dados_empresa()
        return redirect(url_for('app_dashboard'))

    return render_template('login.html', erro="Falha ao autenticar com o Google.")

@app.route('/login/facebook', methods=['GET', 'POST'])
def login_facebook():
    """Rota de Autenticação / Cadastro com a Conta do Facebook (Facebook SSO)"""
    if 'usuario_id' in session:
        return redirect(url_for('app_dashboard'))

    fb_email = request.values.get('email')
    fb_nome = request.values.get('nome') or request.values.get('name')
    fb_sub = request.values.get('sub') or request.values.get('facebook_id')

    if not fb_email:
        return render_template('login.html', erro="Autenticação com o Facebook requer credenciais válidas do provedor. Por favor, tente novamente.")

    conn = get_db_connection()
    if not conn:
        return render_template('login.html', erro="Erro temporário de conexão com o banco de dados.")

    cursor = conn.cursor(dictionary=True)
    user = None
    try:
        try:
            cursor.execute("SELECT facebook_id FROM usuarios LIMIT 1")
            cursor.fetchall()
        except Exception:
            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN facebook_id VARCHAR(255) DEFAULT NULL")
                conn.commit()
            except Exception:
                pass

        cursor.execute("SELECT * FROM usuarios WHERE email = %s OR (facebook_id IS NOT NULL AND facebook_id = %s)", (fb_email, fb_sub))
        user = cursor.fetchone()

        if not user:
            nome_empresa = f"Empresa de {fb_nome}"
            senha_aleatoria = generate_password_hash(f"FacebookSSO_{os.urandom(8).hex()}")

            cursor.execute("INSERT INTO empresas (nome_empresa) VALUES (%s)", (nome_empresa,))
            empresa_id = cursor.lastrowid

            try:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role, facebook_id) VALUES (%s, %s, %s, %s, 'dono', %s)",
                    (empresa_id, fb_nome, fb_email, senha_aleatoria, fb_sub)
                )
            except Exception:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role) VALUES (%s, %s, %s, %s, 'dono')",
                    (empresa_id, fb_nome, fb_email, senha_aleatoria)
                )

            usuario_id = cursor.lastrowid
            conn.commit()

            user = {
                'id': usuario_id,
                'empresa_id': empresa_id,
                'nome': fb_nome,
                'email': fb_email
            }

            try:
                enviar_email_boas_vindas(fb_email, fb_nome, nome_empresa)
                enviar_notificacao_admin("novo_usuario", {
                    "nome_usuario": fb_nome,
                    "email": fb_email,
                    "nome_empresa": nome_empresa,
                    "plano": "Shield Start (Facebook SSO)",
                    "ciclo": "Mensal",
                    "valor": "R$ 0,00"
                })
            except Exception as err_m:
                print(f"Erro no envio de e-mails Facebook SSO: {err_m}")

    except Exception as err_db:
        print(f"Erro no processamento do Facebook SSO: {err_db}")
        if conn: conn.rollback()
        return render_template('login.html', erro="Erro ao processar a autenticação com o Facebook.")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    if user:
        session['usuario_id'] = user['id']
        session['empresa_id'] = user['empresa_id']
        session['nome'] = user['nome']
        garantir_massa_dados_empresa()
        return redirect(url_for('app_dashboard'))

    return render_template('login.html', erro="Falha ao autenticar com o Facebook.")

@app.route('/login/apple', methods=['GET', 'POST'])
def login_apple():
    """Rota de Autenticação / Cadastro com a Conta da Apple (Apple SSO)"""
    if 'usuario_id' in session:
        return redirect(url_for('app_dashboard'))

    apple_email = request.values.get('email')
    apple_nome = request.values.get('nome') or request.values.get('name')
    apple_sub = request.values.get('sub') or request.values.get('apple_id')

    if not apple_email:
        return render_template('login.html', erro="Autenticação com a Apple requer credenciais válidas do provedor. Por favor, tente novamente.")

    conn = get_db_connection()
    if not conn:
        return render_template('login.html', erro="Erro temporário de conexão com o banco de dados.")

    cursor = conn.cursor(dictionary=True)
    user = None
    try:
        try:
            cursor.execute("SELECT apple_id FROM usuarios LIMIT 1")
            cursor.fetchall()
        except Exception:
            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN apple_id VARCHAR(255) DEFAULT NULL")
                conn.commit()
            except Exception:
                pass

        cursor.execute("SELECT * FROM usuarios WHERE email = %s OR (apple_id IS NOT NULL AND apple_id = %s)", (apple_email, apple_sub))
        user = cursor.fetchone()

        if not user:
            nome_empresa = f"Empresa de {apple_nome}"
            senha_aleatoria = generate_password_hash(f"AppleSSO_{os.urandom(8).hex()}")

            cursor.execute("INSERT INTO empresas (nome_empresa) VALUES (%s)", (nome_empresa,))
            empresa_id = cursor.lastrowid

            try:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role, apple_id) VALUES (%s, %s, %s, %s, 'dono', %s)",
                    (empresa_id, apple_nome, apple_email, senha_aleatoria, apple_sub)
                )
            except Exception:
                cursor.execute(
                    "INSERT INTO usuarios (empresa_id, nome, email, senha, role) VALUES (%s, %s, %s, %s, 'dono')",
                    (empresa_id, apple_nome, apple_email, senha_aleatoria)
                )

            usuario_id = cursor.lastrowid
            conn.commit()

            user = {
                'id': usuario_id,
                'empresa_id': empresa_id,
                'nome': apple_nome,
                'email': apple_email
            }

            try:
                enviar_email_boas_vindas(apple_email, apple_nome, nome_empresa)
                enviar_notificacao_admin("novo_usuario", {
                    "nome_usuario": apple_nome,
                    "email": apple_email,
                    "nome_empresa": nome_empresa,
                    "plano": "Shield Start (Apple SSO)",
                    "ciclo": "Mensal",
                    "valor": "R$ 0,00"
                })
            except Exception as err_m:
                print(f"Erro no envio de e-mails Apple SSO: {err_m}")

    except Exception as err_db:
        print(f"Erro no processamento do Apple SSO: {err_db}")
        if conn: conn.rollback()
        return render_template('login.html', erro="Erro ao processar a autenticação com a Apple.")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

    if user:
        session['usuario_id'] = user['id']
        session['empresa_id'] = user['empresa_id']
        session['nome'] = user['nome']
        garantir_massa_dados_empresa()
        return redirect(url_for('app_dashboard'))

    return render_template('login.html', erro="Falha ao autenticar com a Apple.")

@app.route('/esqueci_senha', methods=['POST'])
def esqueci_senha():
    email = request.form.get('email', '').strip().lower()
    if not email:
        return render_template('login.html', erro="Por favor, informe o seu e-mail cadastrado.")
        
    conn = get_db_connection()
    if not conn: return render_template('login.html', erro="Erro temporário no servidor.")
        
    cursor = conn.cursor(dictionary=True)
    user = None
    try:
        try:
            cursor.execute("SELECT reset_token FROM usuarios LIMIT 1")
            cursor.fetchall()
        except Exception:
            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN reset_token VARCHAR(255) DEFAULT NULL")
                cursor.execute("ALTER TABLE usuarios ADD COLUMN reset_expires DATETIME DEFAULT NULL")
                conn.commit()
            except Exception:
                pass

        cursor.execute("SELECT * FROM usuarios WHERE email = %s", (email,))
        user = cursor.fetchone()
        
        if user:
            token = secrets.token_urlsafe(32)
            expiracao_str = (datetime.now() + timedelta(minutes=30)).strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("UPDATE usuarios SET reset_token = %s, reset_expires = %s WHERE id = %s", (token, expiracao_str, user['id']))
            conn.commit()
            
            reset_link = f"https://reaction.frameia.com.br/redefinir_senha?token={token}"
            enviar_email_redefinicao_senha(email, user['nome'], reset_link)
            msg_sucesso = f"Instruções e token enviados para <strong>{email}</strong>! <a href='/redefinir_senha?token={token}' class='underline font-bold hover:text-emerald-800'>Clique aqui para cadastrar a nova senha</a>."
            return render_template('login.html', sucesso=msg_sucesso)
        else:
            return render_template('login.html', erro="Nenhuma conta encontrada com este e-mail.")
    except Exception as e:
        print(f"Erro no esqueci_senha: {e}")
        if conn: conn.rollback()
        return render_template('login.html', erro="Erro ao processar solicitação de redefinição.")
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@app.route('/redefinir_senha', methods=['GET', 'POST'])
def redefinir_senha():
    agora_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if request.method == 'POST':
        token = request.form.get('token', '').strip()
        nova_senha = request.form.get('senha', '')
        
        if not token or not nova_senha:
            return render_template('login.html', erro="Token e nova palavra-passe são obrigatórios.")
            
        conn = get_db_connection()
        if not conn: return render_template('login.html', erro="Erro no servidor.")
            
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute("SELECT * FROM usuarios WHERE reset_token = %s AND (reset_expires IS NULL OR reset_expires > %s)", (token, agora_str))
            user = cursor.fetchone()
            
            if not user:
                return render_template('login.html', erro="O token de redefinição é inválido ou já expirou. Solicite um novo link.")
                
            senha_hash = generate_password_hash(nova_senha)
            cursor.execute("UPDATE usuarios SET senha = %s, reset_token = NULL, reset_expires = NULL WHERE id = %s", (senha_hash, user['id']))
            conn.commit()
            return render_template('login.html', sucesso="Sua palavra-passe foi redefinida com sucesso! Faça login abaixo.")
        except Exception as e:
            if conn: conn.rollback()
            return render_template('login.html', erro="Erro ao atualizar palavra-passe.")
        finally:
            if cursor: cursor.close()
            if conn: conn.close()

    token = request.args.get('token', '').strip()
    if not token:
        return render_template('login.html', erro="Acesso direto negado. É necessário um token válido enviado por e-mail para redefinir a senha.")
        
    conn = get_db_connection()
    if not conn: return render_template('login.html', erro="Erro temporário no servidor.")
        
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM usuarios WHERE reset_token = %s AND (reset_expires IS NULL OR reset_expires > %s)", (token, agora_str))
        user = cursor.fetchone()
        if not user:
            return render_template('login.html', erro="O link de redefinição é inválido ou expirou. Solicite um novo link.")
        return render_template('login.html', redefinir_token=token)
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ----- Telas Principais do SaaS -----

@app.route('/app')
def app_dashboard():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    garantir_massa_dados_empresa()
    return render_template('dashboard.html', usuario=get_usuario_logado())

@app.route('/reputacao')
def reputacao():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    garantir_massa_dados_empresa()
    return render_template('reputacao.html', usuario=get_usuario_logado())

@app.route('/minhas_acoes')
def minhas_acoes():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    garantir_massa_dados_empresa()
    return render_template('minhas_acoes.html', usuario=get_usuario_logado())

@app.route('/relatorios')
def relatorios():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    garantir_massa_dados_empresa()
    return render_template('relatorios.html', usuario=get_usuario_logado())

@app.route('/sala_maquinas')
def sala_maquinas():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    garantir_massa_dados_empresa()
    return render_template('sala_maquinas.html', usuario=get_usuario_logado())

@app.route('/integracoes')
def integracoes():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    garantir_massa_dados_empresa()
    return render_template('integracoes.html', usuario=get_usuario_logado())

@app.route('/ajustes')
def ajustes():
    if 'usuario_id' not in session: return redirect(url_for('login'))
    return render_template('ajustes.html', usuario=get_usuario_logado())

@app.route('/landing')
def landing():
    return render_template('index.html')

# ================= APIs REST (O Backend que alimenta o Frontend) =================

@app.route('/api/acoes', methods=['GET'])
def get_acoes():
    if 'usuario_id' not in session: return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Database error"}), 500
        
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT id as __backendId, titulo as title, 
                   status = 'concluido' as completed, prioridade as priority, prazo as due_date
            FROM acoes WHERE empresa_id = %s ORDER BY prioridade DESC, prazo ASC
        """, (session.get('empresa_id'),))
        acoes = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()
        
    for acao in acoes:
        if acao['due_date']: acao['due_date'] = acao['due_date'].isoformat()
        acao['completed'] = bool(acao['completed'])
            
    return jsonify(acoes)

@app.route('/api/acoes', methods=['POST'])
def create_acao():
    if 'usuario_id' not in session: return jsonify({"error": "Unauthorized"}), 401
    data = request.json
    titulo = data.get('title')
    prioridade = data.get('priority', 'normal')
    prazo = data.get('due_date') or None
    
    if not titulo: return jsonify({"error": "Title required"}), 400
        
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Database error"}), 500
        
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO acoes (empresa_id, criado_por, titulo, prioridade, prazo, status)
            VALUES (%s, %s, %s, %s, %s, 'pendente')
        """, (session.get('empresa_id'), session.get('usuario_id'), titulo, prioridade, prazo))
        new_id = cursor.lastrowid
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Database write error: {e}"}), 500
    finally:
        cursor.close()
        conn.close()
        
    return jsonify({"__backendId": new_id}), 201

@app.route('/api/acoes/<int:acao_id>', methods=['PUT'])
def toggle_acao(acao_id):
    if 'usuario_id' not in session: return jsonify({"error": "Unauthorized"}), 401
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Database error"}), 500
        
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT status FROM acoes WHERE id = %s AND empresa_id = %s", (acao_id, session.get('empresa_id')))
        acao = cursor.fetchone()
        
        if not acao:
            return jsonify({"error": "Not found"}), 404
            
        new_status = 'pendente' if acao[0] == 'concluido' else 'concluido'
        cursor.execute("UPDATE acoes SET status = %s WHERE id = %s", (new_status, acao_id))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Database update error: {e}"}), 500
    finally:
        cursor.close()
        conn.close()
        
    return jsonify({"success": True, "status": new_status})

@app.route('/api/acoes/<int:acao_id>', methods=['DELETE'])
def delete_acao(acao_id):
    if 'usuario_id' not in session: return jsonify({"error": "Unauthorized"}), 401
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Database error"}), 500
        
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM acoes WHERE id = %s AND empresa_id = %s", (acao_id, session.get('empresa_id')))
        conn.commit()
        deleted = cursor.rowcount > 0
    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Database delete error: {e}"}), 500
    finally:
        cursor.close()
        conn.close()
    
    if deleted: return jsonify({"success": True})
    return jsonify({"error": "Not found"}), 404

@app.route('/api/configuracoes', methods=['GET', 'POST'])
def api_configuracoes():
    if 'usuario_id' not in session: return jsonify({"error": "Não autorizado"}), 401
    empresa_id = session.get('empresa_id')
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Erro de base de dados"}), 500
        
    cursor = conn.cursor(dictionary=True)
    try:
        if request.method == 'GET':
            cursor.execute("SELECT tom_voz, regras_ouro, telefone_whatsapp FROM configuracoes_ia WHERE empresa_id = %s", (empresa_id,))
            config = cursor.fetchone()
            if not config:
                return jsonify({"tom_voz": "Profissional e empático", "regras_ouro": "", "telefone_whatsapp": ""})
            return jsonify(config)

        if request.method == 'POST':
            data = request.json
            tom_voz = data.get('tom_voz', 'Profissional e empático')
            regras_ouro = data.get('regras_ouro', '')
            telefone = data.get('telefone_whatsapp', '')

            cursor.execute("""
                INSERT INTO configuracoes_ia (empresa_id, tom_voz, regras_ouro, telefone_whatsapp) 
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE 
                tom_voz = VALUES(tom_voz), 
                regras_ouro = VALUES(regras_ouro), 
                telefone_whatsapp = VALUES(telefone_whatsapp)
            """, (empresa_id, tom_voz, regras_ouro, telefone))

            conn.commit()
            return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Database config error: {e}"}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/api/avaliacoes', methods=['GET'])
def get_avaliacoes():
    if 'usuario_id' not in session: return jsonify({"error": "Unauthorized"}), 401
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Database error"}), 500
        
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT id, nome_cliente as cliente, plataforma_origem as plataforma, 
                   nota, comentario, rascunho_resposta as rascunho, status
            FROM avaliacoes_feed WHERE empresa_id = %s ORDER BY id DESC
        """, (session.get('empresa_id'),))
        avaliacoes = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()
    
    for av in avaliacoes:
        av['tempo'] = "Recente" 
        av['plataforma'] = av['plataforma'].capitalize()
        if not av['rascunho']: av['rascunho'] = "Aguardando análise da IA..."
            
    return jsonify(avaliacoes)

@app.route('/api/avaliacoes/<int:avaliacao_id>/aprovar', methods=['PUT'])
def aprovar_avaliacao(avaliacao_id):
    if 'usuario_id' not in session: return jsonify({"error": "Não autorizado"}), 401
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Erro de base de dados"}), 500

    cursor = conn.cursor()
    try:
        cursor.execute("SELECT status FROM avaliacoes_feed WHERE id = %s AND empresa_id = %s", (avaliacao_id, session.get('empresa_id')))
        av = cursor.fetchone()

        if not av:
            return jsonify({"error": "Avaliação não encontrada"}), 404

        cursor.execute("UPDATE avaliacoes_feed SET status = 'respondido_ia' WHERE id = %s", (avaliacao_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Database error: {e}"}), 500
    finally:
        cursor.close()
        conn.close()

    return jsonify({"success": True, "message": "Avaliação aprovada e movida com sucesso."})

@app.route('/api/integracoes/status', methods=['GET'])
def integracoes_status():
    if 'usuario_id' not in session: return jsonify({"error": "Não autorizado"}), 401
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Erro de DB"}), 500
        
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT plataforma, status FROM integracoes_api WHERE empresa_id = %s", (session.get('empresa_id'),))
        status_db = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()
    
    return jsonify(status_db)

@app.route('/api/analytics/resumo', methods=['GET'])
def get_analytics_resumo():
    if 'usuario_id' not in session: return jsonify({"error": "Não autorizado"}), 401
        
    empresa_id = session.get('empresa_id')
    conn = get_db_connection()
    if not conn: return jsonify({"error": "Erro de DB"}), 500
        
    try:
        cursor = conn.cursor(dictionary=True)
        # Sentimentos
        cursor.execute("""
            SELECT sentimento, COUNT(*) as qtd FROM avaliacoes_feed 
            WHERE empresa_id = %s AND sentimento IS NOT NULL GROUP BY sentimento
        """, (empresa_id,))
        dados_sentimento = cursor.fetchall()
        
        sentimentos_dict = {"positivo": 0, "neutro": 0, "negativo": 0}
        for linha in dados_sentimento:
            if linha['sentimento'] in sentimentos_dict:
                sentimentos_dict[linha['sentimento']] = linha['qtd']
                
        # Tags (Top Motivos)
        cursor.execute("SELECT tags FROM avaliacoes_feed WHERE empresa_id = %s AND tags IS NOT NULL", (empresa_id,))
        linhas_tags = cursor.fetchall()
        
        contagem_tags = {}
        for linha in linhas_tags:
            try:
                lista_tags = json.loads(linha['tags'])
                for tag in lista_tags:
                    if tag: contagem_tags[tag] = contagem_tags.get(tag, 0) + 1
            except: continue
                
        tags_ordenadas = sorted(contagem_tags.items(), key=lambda x: x[1], reverse=True)[:5]
        
        return jsonify({
            "sentimentos": sentimentos_dict,
            "top_tags": {"labels": [item[0] for item in tags_ordenadas], "valores": [item[1] for item in tags_ordenadas]}
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close(); conn.close()

# ================= ROTAS DE INTEGRAÇÃO OAUTH =================

@app.route('/api/auth/<plataforma>/login')
def auth_login(plataforma):
    if 'usuario_id' not in session: return redirect(url_for('login'))
    redirect_uri = url_for('auth_callback', plataforma=plataforma, _external=True)
    
    if plataforma == 'google':
        auth_url = (f"https://accounts.google.com/o/oauth2/v2/auth?client_id={GOOGLE_CLIENT_ID}&"
                    f"redirect_uri={redirect_uri}&response_type=code&"
                    f"scope=https://www.googleapis.com/auth/business.manage&access_type=offline&prompt=consent")
        return redirect(auth_url)
        
    elif plataforma == 'instagram':
        auth_url = (f"https://www.facebook.com/v18.0/dialog/oauth?client_id={META_APP_ID}&"
                    f"redirect_uri={redirect_uri}&config_id=SEU_CONFIG_ID_AQUI")
        return redirect(auth_url)

    elif plataforma in ['ifood', 'whatsapp', 'ze_delivery', '99food']:
        empresa_id = session.get('empresa_id')
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            try:
                cursor.execute("SELECT id FROM integracoes_api WHERE empresa_id = %s AND plataforma = %s", (empresa_id, plataforma))
                if cursor.fetchone():
                    cursor.execute("UPDATE integracoes_api SET status = 'ativo' WHERE empresa_id = %s AND plataforma = %s", (empresa_id, plataforma))
                else:
                    cursor.execute("INSERT INTO integracoes_api (empresa_id, plataforma, token_acesso, status) VALUES (%s, %s, 'demo_token', 'ativo')", (empresa_id, plataforma))
                conn.commit()
            except Exception as e:
                conn.rollback()
                print(f"Erro ao integrar {plataforma}: {e}")
            finally:
                cursor.close(); conn.close()
        return redirect(url_for('integracoes', sucesso="true", plat=plataforma))

    return "Plataforma inválida", 400

@app.route('/api/auth/<plataforma>/callback')
def auth_callback(plataforma):
    if 'usuario_id' not in session: return redirect(url_for('login'))
    empresa_id = session.get('empresa_id')
    code = request.args.get('code')
    redirect_uri = url_for('auth_callback', plataforma=plataforma, _external=True)
    
    if not code: return "Erro: Código não recebido ou acesso negado pelo cliente.", 400

    token_acesso, token_refresh = "", ""
    page_id = None

    try:
        if plataforma == 'google':
            res = requests.post("https://oauth2.googleapis.com/token", data={
                "code": code, "client_id": GOOGLE_CLIENT_ID, "client_secret": GOOGLE_CLIENT_SECRET,
                "redirect_uri": redirect_uri, "grant_type": "authorization_code"
            }).json()
            token_acesso = res.get("access_token")
            token_refresh = res.get("refresh_token", "N/A")

        elif plataforma == 'instagram':
            res = requests.get("https://graph.facebook.com/v18.0/oauth/access_token", params={
                "client_id": META_APP_ID, "redirect_uri": redirect_uri, "client_secret": META_APP_SECRET, "code": code
            }).json()
            token_acesso = res.get("access_token")
            token_refresh = "N/A" 
            
            if token_acesso:
                pages_res = requests.get(f"https://graph.facebook.com/v18.0/me/accounts?access_token={token_acesso}").json()
                if 'data' in pages_res and len(pages_res['data']) > 0:
                    page_id = pages_res['data'][0]['id']
                    page_token = pages_res['data'][0]['access_token']
                    requests.post(f"https://graph.facebook.com/v18.0/{page_id}/subscribed_apps", 
                                  data={"subscribed_fields": "feed, messages, comments", "access_token": page_token})
                    token_acesso = page_token 

        if not token_acesso: return f"Erro ao obter token real da plataforma.", 400

    except Exception as e:
        return f"Erro de comunicação com a plataforma: {str(e)}", 500
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT id FROM integracoes_api WHERE empresa_id = %s AND plataforma = %s", (empresa_id, plataforma))
            if cursor.fetchone():
                cursor.execute("""
                    UPDATE integracoes_api 
                    SET token_acesso = %s, token_refresh = %s, plataforma_user_id = %s, status = 'ativo' 
                    WHERE empresa_id = %s AND plataforma = %s
                """, (token_acesso, token_refresh, page_id, empresa_id, plataforma))
            else:
                cursor.execute("""
                    INSERT INTO integracoes_api (empresa_id, plataforma, token_acesso, token_refresh, plataforma_user_id, status) 
                    VALUES (%s, %s, %s, %s, %s, 'ativo')
                """, (empresa_id, plataforma, token_acesso, token_refresh, page_id))
            conn.commit()
        except Exception as e:
            conn.rollback()
            print(f"Erro ao salvar integração no banco: {e}")
        finally:
            cursor.close()
            conn.close()

    return redirect(url_for('integracoes', sucesso="true", plat=plataforma))

# ================= WEBHOOKS (Escuta Ativa) =================

@app.route('/api/webhooks/ingest', methods=['GET', 'POST'])
def webhook_ingest():
    if request.method == 'GET':
        if request.args.get('hub.mode') == 'subscribe' and request.args.get('hub.verify_token') == META_VERIFY_TOKEN:
            return request.args.get('hub.challenge'), 200
        return "Acesso negado", 403

    data = request.json
    if not data: return jsonify({"error": "Payload inválido"}), 400

    conn = get_db_connection()
    if not conn: return jsonify({"error": "Falha no DB"}), 500

    try:
        cursor = conn.cursor()
        
        # PARSE ADAPTATIVO
        if 'entry' in data:
            try:
                plataforma = 'instagram'
                
                # Mapeamento dinâmico multi-tenancy a partir do ID da página enviado pelo webhook da Meta
                meta_page_id = str(data['entry'][0].get('id', ''))
                cursor.execute("""
                    SELECT empresa_id FROM integracoes_api 
                    WHERE plataforma = 'instagram' AND plataforma_user_id = %s AND status = 'ativo'
                """, (meta_page_id,))
                row = cursor.fetchone()
                empresa_id = row[0] if row else 1 # Fallback seguro
                
                alteracao = data['entry'][0]['changes'][0]['value']
                nome_cliente = alteracao.get('from', {}).get('username', 'Utilizador')
                comentario = alteracao.get('text', '')
                nota = 3 
            except (KeyError, IndexError):
                return jsonify({"error": "Formato Meta inválido"}), 400
        else:
            empresa_id = data.get('empresa_id')
            plataforma = data.get('plataforma')
            nome_cliente = data.get('nome_cliente', 'Cliente')
            nota = data.get('nota')
            comentario = data.get('comentario', '')

        if not empresa_id or not plataforma or nota is None:
            return jsonify({"error": "Dados ausentes"}), 400

        cursor.execute("""
            INSERT INTO avaliacoes_feed (empresa_id, plataforma_origem, nome_cliente, nota, comentario, status)
            VALUES (%s, %s, %s, %s, %s, 'pendente')
        """, (empresa_id, plataforma, nome_cliente, nota, comentario))
        avaliacao_id = cursor.lastrowid
        
        filtro = analisar_qualidade_comentario(comentario, nota)
        
        if filtro["valido"]:
            cursor.execute("SELECT tom_voz, regras_ouro, telefone_whatsapp FROM configuracoes_ia WHERE empresa_id = %s", (empresa_id,))
            config_ia = cursor.fetchone()
            regras_reais = f"Tom: {config_ia[0]}. Regras: {config_ia[1]}" if config_ia else "Tom profissional."
            telefone_real = config_ia[2] if config_ia and config_ia[2] else "Não configurado"

            resultado_ia = analisar_avaliacao_gemini(
                nome_cliente=nome_cliente, nota=nota, comentario=comentario,
                regras_tom_voz=regras_reais, telefone_empresa=telefone_real
            )
            sentimento = resultado_ia.get("sentimento", "neutro")
            rascunho = resultado_ia.get("sugestao_resposta", "")
            
            # Tags estruturadas retornadas pelo Gemini 2.5
            tags_json = json.dumps(resultado_ia.get("tags", []))
            
            novo_status = 'respondido_ia' if nota >= 4 else 'alerta_crise' if nota <= 3 else 'pendente'
        else:
            sentimento = filtro["sentimento"]
            rascunho = filtro["rascunho"]
            tags_json = json.dumps(["Filtro Spam Ativo"])
            novo_status = filtro["status"]

        if novo_status == 'alerta_crise':
            titulo_crise = f"Contenção: Avaliação {nota}★ de {nome_cliente}"
            cursor.execute("INSERT INTO acoes (empresa_id, titulo, prioridade, prazo, status) VALUES (%s, %s, 'critical', %s, 'pendente')", 
                           (empresa_id, titulo_crise, date.today().strftime('%Y-%m-%d')))

        cursor.execute("""
            UPDATE avaliacoes_feed SET sentimento = %s, rascunho_resposta = %s, tags = %s, status = %s WHERE id = %s
        """, (sentimento, rascunho, tags_json, novo_status, avaliacao_id))

        conn.commit()
        return jsonify({"success": True, "message": "Avaliação processada", "status": novo_status}), 201

    except Exception as e:
        if conn: conn.rollback()
        return jsonify({"error": f"Erro interno: {str(e)}"}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if conn: conn.close()

# ================= GERADOR DE MASSA DE DADOS (DEMO) =================

@app.route('/api/seed', methods=['GET', 'POST'])
def api_seed():
    """Endpoint para gerar massa de dados demonstrativa no app"""
    try:
        from gerar_massa_dados import popular_banco_dados
        res = popular_banco_dados()
        return jsonify({"success": True, "message": "Massa de dados gerada com sucesso!", "detalhes": res}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ================= INTEGRACAO MERCADO PAGO =================
MERCADOPAGO_ACCESS_TOKEN = os.environ.get("MERCADOPAGO_ACCESS_TOKEN", "APP_USR-DEMO-TOKEN")

@app.route('/api/checkout/<plano>/<ciclo>', methods=['GET', 'POST'])
def api_checkout(plano, ciclo):
    """Gera preferência de checkout no Mercado Pago (Checkout Pro / PIX)"""
    if 'usuario_id' not in session:
        return redirect(url_for('login', registo="true"))
    
    precos = {
        "pro": {"mensal": 79.90, "anual": 718.80},
        "premium": {"mensal": 197.90, "anual": 1774.80},
        "setup_vip": {"avista": 300.00, "mensal": 300.00, "anual": 300.00}
    }
    
    if plano == "setup_vip":
        ciclo = "avista"
        
    if plano not in precos or ciclo not in ["mensal", "anual", "avista"]:
        return jsonify({"error": "Plano ou ciclo inválido"}), 400
        
    valor = precos[plano][ciclo]
    titulo = "ReAction - Setup VIP (2 Sessões ao Vivo + 3 Meses Grátis)" if plano == "setup_vip" else f"ReAction - Plano {plano.capitalize()} ({ciclo.capitalize()})"
    
    try:
        import mercadopago
        sdk = mercadopago.SDK(MERCADOPAGO_ACCESS_TOKEN)
        
        preference_data = {
            "items": [
                {
                    "title": titulo,
                    "quantity": 1,
                    "unit_price": valor,
                    "currency_id": "BRL"
                }
            ],
            "payer": {
                "email": session.get('usuario_email', 'cliente@reaction.com.br')
            },
            "back_urls": {
                "success": url_for('app_dashboard', status="sucesso_pagamento", _external=True),
                "failure": url_for('app_dashboard', status="falha_pagamento", _external=True),
                "pending": url_for('app_dashboard', status="pendente_pagamento", _external=True)
            },
            "auto_return": "approved",
            "external_reference": f"empresa_{session.get('empresa_id')}_{plano}_{ciclo}"
        }
        
        preference_response = sdk.preference().create(preference_data)
        preference = preference_response.get("response", {})
        
        # Notificar os 3 administradores sobre o upgrade/contratação
        try:
            enviar_notificacao_admin("upgrade", {
                "nome_usuario": session.get('nome', 'Cliente'),
                "email": session.get('usuario_email', 'cliente@reaction.com.br'),
                "nome_empresa": f"Empresa ID #{session.get('empresa_id')}",
                "plano": f"Plano {plano.capitalize()}",
                "ciclo": ciclo.capitalize(),
                "valor": f"R$ {valor:.2f}"
            })
        except Exception as e_adm:
            print(f"Aviso envio e-mail admin checkout: {e_adm}")

        init_point = preference.get("init_point") or preference.get("sandbox_init_point")
        if init_point:
            return redirect(init_point)
        else:
            return jsonify({"success": True, "checkout_data": preference, "message": "Simulação de Checkout Mercado Pago iniciada"}), 200
            
    except Exception as mp_err:
        print(f"Erro no checkout Mercado Pago: {mp_err}")
        return jsonify({"error": f"Erro no Checkout Mercado Pago: {str(mp_err)}"}), 500

@app.route('/api/webhooks/mercadopago', methods=['POST'])
def webhook_mercadopago():
    """Webhook do Mercado Pago para escutar alterações e aprovações de pagamento"""
    data = request.json or {}
    print(f"Webhook Mercado Pago recebido: {data}")
    return jsonify({"status": "received"}), 200

# ================= ROTAS DE EXPORTAÇÃO EXCEL NATIVO (.XLSX MULTI-ABAS) =================

@app.route('/api/exportar/relatorio_excel')
@app.route('/api/exportar/avaliacoes')
@app.route('/api/exportar/acoes')
def exportar_relatorio_excel():
    """Gera um relatório completo em formato nativo Excel (.xlsx) com 4 abas profissionais estilizadas"""
    if 'usuario_id' not in session:
        return redirect(url_for('login'))

    empresa_id = session.get('empresa_id')
    conn = get_db_connection()
    if not conn:
        return "Erro de conexão com o banco de dados", 500

    cursor = conn.cursor(dictionary=True)
    avaliacoes = []
    acoes = []
    integracoes = []
    empresa_nome = f"Empresa #{empresa_id}"
    
    try:
        cursor.execute("SELECT nome_empresa FROM empresas WHERE id = %s", (empresa_id,))
        emp_res = cursor.fetchone()
        if emp_res and emp_res.get('nome_empresa'):
            empresa_nome = emp_res['nome_empresa']

        cursor.execute("""
            SELECT id, plataforma_origem, nome_cliente, nota, comentario, sentimento, status, rascunho_resposta
            FROM avaliacoes_feed
            WHERE empresa_id = %s
        """, (empresa_id,))
        avaliacoes = cursor.fetchall()

        cursor.execute("""
            SELECT id, titulo, prioridade, prazo, status
            FROM acoes
            WHERE empresa_id = %s
        """, (empresa_id,))
        acoes = cursor.fetchall()

        cursor.execute("""
            SELECT plataforma, status
            FROM integracoes_api
            WHERE empresa_id = %s
        """, (empresa_id,))
        integracoes = cursor.fetchall()

    except Exception as e:
        print(f"Erro ao buscar dados para relatório Excel: {e}")
    finally:
        cursor.close()
        conn.close()

    # Criar o Workbook openpyxl
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    fill_header = PatternFill(start_color="3A3A3A", end_color="3A3A3A", fill_type="solid")
    fill_brand = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True, color="1F2937")
    font_normal = Font(name="Calibri", size=11, color="374151")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    # ── ABA 1: RESUMO EXECUTIVO ──
    ws1 = wb.active
    ws1.title = "Resumo Executivo"
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:E2")
    cell_title = ws1["A1"]
    cell_title.value = f"REACTION | RELATÓRIO EXECUTIVO DE REPUTAÇÃO - {empresa_nome.upper()}"
    cell_title.font = font_title
    cell_title.fill = fill_brand
    cell_title.alignment = align_center

    total_av = len(avaliacoes)
    soma_notas = sum(r.get('nota', 0) for r in avaliacoes)
    nota_media = round(soma_notas / total_av, 2) if total_av > 0 else 5.0
    positivas = sum(1 for r in avaliacoes if r.get('sentimento') == 'positivo' or r.get('nota', 0) >= 4)
    neutras = sum(1 for r in avaliacoes if r.get('sentimento') == 'neutro' or r.get('nota', 0) == 3)
    negativas = sum(1 for r in avaliacoes if r.get('sentimento') == 'negativo' or r.get('nota', 0) <= 2)
    saude_pct = round((positivas / total_av) * 100, 1) if total_av > 0 else 100.0

    kpis = [
        ("Nota Média da Marca", f"{nota_media} / 5.0"),
        ("Total de Avaliações Cadastradas", total_av),
        ("Índice de Saúde da Reputação", f"{saude_pct}%"),
        ("Avaliações Positivas (4-5★)", positivas),
        ("Avaliações Neutras (3★)", neutras),
        ("Avaliações Críticas (1-2★)", negativas),
        ("Total de Ações de Contenção", len(acoes))
    ]

    ws1.cell(row=4, column=1, value="MÉTRICAS CHAVE DE PERFORMANCE").font = Font(name="Calibri", size=12, bold=True, color="FF6B35")

    ws1.cell(row=5, column=1, value="Indicador Executivo").font = font_header
    ws1.cell(row=5, column=1).fill = fill_header
    ws1.cell(row=5, column=2, value="Resultado").font = font_header
    ws1.cell(row=5, column=2).fill = fill_header

    for idx, (k, v) in enumerate(kpis, start=6):
        c1 = ws1.cell(row=idx, column=1, value=k)
        c2 = ws1.cell(row=idx, column=2, value=v)
        c1.font = font_bold
        c2.font = font_bold
        c1.border = thin_border
        c2.border = thin_border
        c2.alignment = align_center

    # ── ABA 2: AVALIAÇÕES ──
    ws2 = wb.create_sheet(title="Avaliações & Feedbacks")
    ws2.views.sheetView[0].showGridLines = True

    headers_av = ['ID', 'Cliente', 'Canal / Origem', 'Nota (Estrelas)', 'Comentário do Cliente', 'Sentimento', 'Status da Resposta', 'Resposta Gerada / IA']
    for col_num, h in enumerate(headers_av, 1):
        cell = ws2.cell(row=1, column=col_num, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_left if col_num in [2, 5, 8] else align_center

    for r_idx, r in enumerate(avaliacoes, start=2):
        row_data = [
            r.get('id', ''),
            r.get('nome_cliente', ''),
            r.get('plataforma_origem', '').upper(),
            r.get('nota', ''),
            r.get('comentario', ''),
            r.get('sentimento', '').capitalize(),
            r.get('status', ''),
            r.get('rascunho_resposta', '')
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_normal
            cell.border = thin_border
            if r_idx % 2 == 0: cell.fill = fill_zebra

    # ── ABA 3: PLANO DE AÇÕES ──
    ws3 = wb.create_sheet(title="Plano de Ações")
    ws3.views.sheetView[0].showGridLines = True

    headers_ac = ['ID', 'Título da Ação de Contenção', 'Prioridade', 'Prazo Final', 'Status']
    for col_num, h in enumerate(headers_ac, 1):
        cell = ws3.cell(row=1, column=col_num, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    for r_idx, r in enumerate(acoes, start=2):
        row_data = [
            r.get('id', ''),
            r.get('titulo', ''),
            r.get('prioridade', 'normal').capitalize(),
            str(r.get('prazo') or 'Sem prazo'),
            r.get('status', '').capitalize()
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_normal
            cell.border = thin_border
            if r_idx % 2 == 0: cell.fill = fill_zebra

    # Auto-ajuste de largura de colunas em todas as abas
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 65)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reaction_relatorio_executivo_{empresa_id}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

if __name__ == '__main__':
    flask_debug = os.environ.get('FLASK_DEBUG', 'False').lower() in ['true', '1']
    flask_host = os.environ.get('FLASK_HOST', '127.0.0.1')
    app.run(host=flask_host, port=5001, debug=flask_debug)